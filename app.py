import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from pyxirr import xirr
from datetime import datetime, timedelta
import requests
import io
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as xl_numbers)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# GOOGLE DRIVE — carga de archivos via Service Account
# ─────────────────────────────────────────────────────────────────────────────
DRIVE_FILE_IDS = {
    "datos":        "1g8GYj_zU0BmfuRLf-Ts6WV2Eq6YFCWJw",
    "curvas_cobalt": "1pj4RInTUAFB-dlNQfgq02-xlPwfgGVmZ",
}

@st.cache_resource(show_spinner=False)
def _get_drive_service():
    """Crea cliente de Google Drive usando las credenciales del secrets.toml."""
    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
        creds_dict = dict(st.secrets["gcp_service_account"])
        creds_dict["private_key"] = creds_dict["private_key"].replace("\\n", "\n")
        creds = service_account.Credentials.from_service_account_info(
            creds_dict, scopes=["https://www.googleapis.com/auth/drive.readonly"]
        )
        return build("drive", "v3", credentials=creds)
    except Exception as e:
        return None

def load_excel_from_drive(file_key: str) -> io.BytesIO:
    """Descarga un Excel de Drive y lo devuelve como BytesIO."""
    file_id = DRIVE_FILE_IDS[file_key]
    service = _get_drive_service()
    if service is None:
        raise RuntimeError("No se pudo conectar a Google Drive. Revisa los secrets.")
    from googleapiclient.http import MediaIoBaseDownload
    request = service.files().get_media(fileId=file_id)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    buf.seek(0)
    return buf

# ─────────────────────────────────────────────────────────────────────────────
# HELPER — genera Excel formateado y devuelve bytes para st.download_button
# ─────────────────────────────────────────────────────────────────────────────
def df_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Data",
                      title: str = "", report_curr: str = "USD") -> bytes:
    """
    Convierte un DataFrame en un Excel formateado estilo Family Office:
    - Header azul oscuro, texto blanco, negrita
    - Filas alternadas gris muy claro / blanco
    - Columnas de dinero con formato #,##0
    - Columnas % con formato 0.00%
    - Columnas x (múltiplos) con formato 0.00x
    - Anchos automáticos
    - Fila de título si se pasa `title`
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    # ── Paleta ────────────────────────────────────────────────────────────────
    HDR_FILL  = PatternFill("solid", fgColor="002060")   # azul oscuro
    ALT_FILL  = PatternFill("solid", fgColor="EEF2FA")   # gris muy claro
    TTL_FILL  = PatternFill("solid", fgColor="0D1929")   # azul marino (título)
    WHT_FONT  = Font(name="Arial", color="FFFFFF", bold=True, size=10)
    HDR_FONT  = Font(name="Arial", color="FFFFFF", bold=True, size=10)
    BODY_FONT = Font(name="Arial", color="1A1A2E", size=10)
    TTL_FONT  = Font(name="Arial", color="FFFFFF", bold=True, size=12)
    THIN      = Side(style="thin", color="D0D8E8")
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    HDR_BRD   = Border(left=THIN, right=THIN,
                        top=Side(style="medium", color="002060"),
                        bottom=Side(style="medium", color="002060"))

    curr_sym  = {"USD": "$", "EUR": "€", "GBP": "£"}.get(report_curr, "$")
    n_cols    = len(df.columns) + 1   # +1 por el índice

    row_offset = 1

    # ── Fila de título ────────────────────────────────────────────────────────
    if title:
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=n_cols)
        cell = ws.cell(row=1, column=1, value=title)
        cell.font      = TTL_FONT
        cell.fill      = TTL_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center",
                                   indent=1)
        ws.row_dimensions[1].height = 22
        row_offset = 2

    # ── Header ────────────────────────────────────────────────────────────────
    hdr_row = row_offset
    ws.cell(row=hdr_row, column=1, value="#").font   = HDR_FONT
    ws.cell(row=hdr_row, column=1).fill              = HDR_FILL
    ws.cell(row=hdr_row, column=1).alignment         = Alignment(horizontal="center")
    ws.cell(row=hdr_row, column=1).border            = HDR_BRD
    ws.row_dimensions[hdr_row].height = 18

    for ci, col in enumerate(df.columns, start=2):
        c = ws.cell(row=hdr_row, column=ci, value=str(col))
        c.font      = HDR_FONT
        c.fill      = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border    = HDR_BRD

    # ── Detectar tipo de columna para formato ─────────────────────────────────
    def col_fmt(col_name: str):
        cn = str(col_name).lower()
        # Cualquier columna con % en el nombre → formato porcentaje literal
        if cn.startswith('%') or ' %' in cn or cn.endswith('%') or 'irr' in cn or 'rent' in cn:
            return '0.00"%"', "right"
        if any(x in cn for x in ["tvpi", "dpi", "rvpi", "múltiplo", "multiple"]):
            return '0.00"x"', "right"
        if any(x in cn for x in ["commit", "paid", "unfunded", "distributed",
                                   "nav", "total value", "utilidad", "ganancia",
                                   "impuesto", "calls", "distribuciones",
                                   "net cash", "value", "(mm)"]):
            return "#,##0.0", "right"
        if cn in ["vintage", "n° fondos", "fondos", "año", "# inv", "duration"]:
            return "0.0", "center"
        return "General", "left"

    # ── Filas de datos ────────────────────────────────────────────────────────
    for ri, (idx, row) in enumerate(df.iterrows(), start=1):
        xr = hdr_row + ri
        fill = ALT_FILL if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

        # Índice
        c = ws.cell(row=xr, column=1, value=idx)
        c.font = BODY_FONT; c.fill = fill; c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")

        for ci, col in enumerate(df.columns, start=2):
            val = row[col]
            fmt, align = col_fmt(col)

            # Limpiar valores formateados como string (ej. "1.33x", "12.58%")
            if isinstance(val, str):
                clean = val.replace("x","").replace("%","").replace(",","").strip()
                try:
                    val = float(clean)
                    # NO dividir por 100 — los IRR ya vienen en escala correcta
                    # (ej. 12.58 significa 12.58%, el formato '0.00"%"' lo muestra bien)
                except ValueError:
                    pass  # dejar como string

            c = ws.cell(row=xr, column=ci, value=val)
            c.font      = BODY_FONT
            c.fill      = fill
            c.border    = BORDER
            c.alignment = Alignment(horizontal=align, vertical="center")
            if fmt != "General":
                c.number_format = fmt

        ws.row_dimensions[xr].height = 16

    # ── Anchos automáticos ────────────────────────────────────────────────────
    for ci in range(1, n_cols + 1):
        max_len = 0
        col_letter = get_column_letter(ci)
        for row_cells in ws.iter_rows(min_col=ci, max_col=ci):
            for cell in row_cells:
                try:
                    cell_len = len(str(cell.value)) if cell.value else 0
                    max_len  = max(max_len, cell_len)
                except:
                    pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 30)

    # ── Freeze panes (fijar header) ───────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=2)

    # ── Guardar en bytes ──────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def excel_download_btn(df: pd.DataFrame, label: str, filename: str,
                       sheet_name: str, title: str, report_curr: str,
                       key: str):
    """Botón de descarga compacto con ícono, alineado a la derecha."""
    xlsx_bytes = df_to_excel_bytes(df, sheet_name=sheet_name,
                                    title=title, report_curr=report_curr)
    # Columna vacía + botón para alinearlo a la derecha
    _, btn_col = st.columns([6, 1])
    with btn_col:
        st.download_button(
            label="⬇ .xlsx",
            data=xlsx_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=key,
            help=f"Descargar {label} en Excel",
            use_container_width=True,
        )


# ─────────────────────────────────────────────────────────────────────────────
# 1. CONFIGURACIÓN E INTERFAZ
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Family Office OS", layout="wide", initial_sidebar_state="expanded")
st.title("🏛️ Alternative Assets Monitor")
st.markdown("---")

# ─────────────────────────────────────────────────────────────────────────────
# CSS — solo toca el sidebar para darle el look CobaltLP
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600&display=swap');

/* ── Sidebar base ── */
[data-testid="stSidebar"] {
    background-color: #0d1929 !important;
    border-right: 1px solid #1a2e45 !important;
    font-family: 'Inter', sans-serif !important;
}
[data-testid="stSidebar"] > div:first-child {
    padding: 0 !important;
}

/* ── Ocultar labels nativos de Streamlit dentro del sidebar ── */
[data-testid="stSidebar"] .stSelectbox label,
[data-testid="stSidebar"] .stMultiSelect label,
[data-testid="stSidebar"] .stCheckbox label span,
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3 {
    color: #3a6080 !important;
    font-size: 13px !important;
    text-transform: uppercase !important;
    letter-spacing: 0.08em !important;
    font-weight: 600 !important;
}

/* ── Selectbox y multiselect dentro del sidebar ── */
[data-testid="stSidebar"] .stSelectbox > div > div,
[data-testid="stSidebar"] .stMultiSelect > div > div {
    background-color: #0a1622 !important;
    border: 1px solid #1a3350 !important;
    border-radius: 5px !important;
    color: #8ab0cc !important;
    font-size: 14px !important;
}

/* Tags del multiselect (las pastillas) */
[data-testid="stSidebar"] .stMultiSelect span[data-baseweb="tag"] {
    background-color: #0e2a45 !important;
    border: 1px solid #1a4a70 !important;
    color: #5baee8 !important;
    font-size: 13px !important;
    border-radius: 3px !important;
}

/* ── Checkbox ── */
[data-testid="stSidebar"] .stCheckbox > label {
    color: #4a6d8c !important;
    font-size: 14px !important;
}

/* ── Botones en sidebar ── */
[data-testid="stSidebar"] .stButton > button {
    background-color: #1a5fd4 !important;
    color: #ffffff !important;
    border: none !important;
    border-radius: 5px !important;
    font-size: 13px !important;
    font-weight: 600 !important;
    letter-spacing: 0.04em !important;
    width: 100% !important;
    padding: 8px 0 !important;
    margin-top: 4px !important;
    transition: background 0.15s !important;
}
[data-testid="stSidebar"] .stButton > button:hover {
    background-color: #1450b0 !important;
}

/* ── Info box (tasas BCE) ── */
[data-testid="stSidebar"] .stAlert {
    background-color: #0a1622 !important;
    border: 1px solid #1a3350 !important;
    border-radius: 5px !important;
    color: #5baee8 !important;
    font-size: 13px !important;
}

/* ── Separador ── */
[data-testid="stSidebar"] hr {
    border-color: #1a2e45 !important;
    margin: 8px 0 !important;
}

/* ── Scrollbar del sidebar ── */
[data-testid="stSidebar"] ::-webkit-scrollbar       { width: 4px; }
[data-testid="stSidebar"] ::-webkit-scrollbar-track { background: #0a1118; }
[data-testid="stSidebar"] ::-webkit-scrollbar-thumb { background: #1a3350; border-radius: 2px; }

/* ── Bloque de header del sidebar (logo) ── */
.sb-logo {
    padding: 16px 16px 14px;
    border-bottom: 1px solid #1a2e45;
    margin-bottom: 4px;
    display: flex; align-items: center; gap: 10px;
}
.sb-logo-mark {
    width: 32px; height: 32px;
    background: #1a5fd4;
    border-radius: 6px;
    display: flex; align-items: center; justify-content: center;
    font-size: 14px; font-weight: 700; color: #fff;
    flex-shrink: 0;
}
.sb-logo-name  { font-size: 15px; font-weight: 600; color: #d4e8f8; letter-spacing: -0.2px; }
.sb-logo-sub   { font-size: 10px; color: #2a4a6a; letter-spacing: 0.1em; text-transform: uppercase; }

/* ── Sección de filtros ── */
.sb-section-label {
    padding: 10px 16px 4px;
    font-size: 11px; font-weight: 700;
    color: #2a4560;
    text-transform: uppercase; letter-spacing: 0.12em;
    border-top: 1px solid #1a2e45;
    margin-top: 6px;
}

/* ── Fila de configuración (moneda / fecha) ── */
.sb-config-label {
    font-size: 11px; font-weight: 600;
    color: #2a4560;
    text-transform: uppercase; letter-spacing: 0.1em;
    margin-bottom: 2px; padding: 0 2px;
}

/* ── Tasas BCE pill ── */
.sb-rates {
    background: #0a1622;
    border: 1px solid #1a3350;
    border-radius: 5px;
    padding: 10px 14px;
    margin: 4px 0 8px;
    font-size: 13px;
}
.sb-rates-title {
    font-size: 11px; font-weight: 700;
    color: #2a4560; letter-spacing: 0.1em;
    text-transform: uppercase; margin-bottom: 6px;
}
.sb-rate-row { display: flex; justify-content: space-between; margin-bottom: 3px; }
.sb-rate-key { color: #3a6080; }
.sb-rate-val { color: #5baee8; font-weight: 600; font-family: monospace; }

/* ── Expander de filtros (CobaltLP style) ── */
[data-testid="stSidebar"] .streamlit-expanderHeader {
    background-color: #0c1e30 !important;
    border: none !important;
    border-bottom: 1px solid #1a2e45 !important;
    border-radius: 0 !important;
    color: #ffffff !important;
    font-size: 14px !important;
    font-weight: 500 !important;
    padding: 10px 16px !important;
    letter-spacing: 0.02em !important;
}
[data-testid="stSidebar"] .streamlit-expanderHeader:hover {
    background-color: #0e2235 !important;
    color: #ffffff !important;
}
[data-testid="stSidebar"] .streamlit-expanderHeader svg {
    color: #5baee8 !important;
}
[data-testid="stSidebar"] .streamlit-expanderContent {
    background-color: #091520 !important;
    border: none !important;
    border-bottom: 1px solid #1a2e45 !important;
    border-radius: 0 !important;
    padding: 12px 14px !important;
}

/* Forzar blanco en TODOS los selectores posibles del expander */
[data-testid="stSidebar"] details > summary,
[data-testid="stSidebar"] details > summary p,
[data-testid="stSidebar"] details > summary span,
[data-testid="stSidebar"] details > summary div,
[data-testid="stSidebar"] details summary *,
[data-testid="stSidebar"] [data-testid="stExpander"] summary,
[data-testid="stSidebar"] [data-testid="stExpander"] summary p,
[data-testid="stSidebar"] [data-testid="stExpander"] summary span {
    color: #ffffff !important;
    font-size: 14px !important;
    font-weight: 500 !important;
}
[data-testid="stSidebar"] details > summary:hover,
[data-testid="stSidebar"] details > summary:hover * {
    color: #ffffff !important;
    background-color: #0e2235 !important;
}

/* ── Botón descarga Excel compacto ── */
.stDownloadButton > button {
    background-color: transparent !important;
    border: 1px solid #d0d8e8 !important;
    border-radius: 5px !important;
    color: #5a7fa0 !important;
    font-size: 11px !important;
    font-weight: 500 !important;
    padding: 3px 10px !important;
    height: 28px !important;
    line-height: 1 !important;
    transition: all 0.15s !important;
}
.stDownloadButton > button:hover {
    background-color: #f0f4fa !important;
    border-color: #1a5fd4 !important;
    color: #1a5fd4 !important;
}

/* ── Botón Clear All del sidebar (no tocar con el estilo de arriba) ── */
.sb-clear-btn {
    margin: 10px 16px 6px;
}
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# 2. MOTOR DE DIVISAS (sin cambios)
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(ttl=3600)
def get_fx_map_institutional(start_date, end_date):
    FALLBACK = {"USD": 1.08, "GBP": 0.86}
    try:
        # Limitar end_date a hoy para evitar errores con fechas futuras
        today = datetime.today().date()
        end_capped = min(end_date.date() if hasattr(end_date,'date') else end_date, today)
        url = f"https://api.frankfurter.app/{start_date.strftime('%Y-%m-%d')}..{end_capped.strftime('%Y-%m-%d')}?to=USD,GBP"
        response = requests.get(url, timeout=15)
        data = response.json()
        rates_raw = data.get('rates', {})
        if not rates_raw:
            raise ValueError("Empty rates")
        fx_dict = {}
        for date_str, v in rates_raw.items():
            fx_dict[datetime.strptime(date_str, '%Y-%m-%d').date()] = {
                "USD": float(v.get("USD", FALLBACK["USD"])),
                "GBP": float(v.get("GBP", FALLBACK["GBP"])),
            }
        full_range = pd.date_range(start_date, end_date)
        clean_map = {}
        last_v = FALLBACK.copy()
        for d in full_range:
            d_d = d.date()
            if d_d in fx_dict:
                last_v = fx_dict[d_d]
            clean_map[d_d] = dict(last_v)
        return clean_map
    except:
        # Si la API falla, devolver mapa con tasas de fallback para todo el rango
        full_range = pd.date_range(start_date, end_date)
        return {d.date(): FALLBACK.copy() for d in full_range}


def convert_amount(amount, from_curr, to_curr, fx_day):
    if from_curr == to_curr:
        return float(amount)
    rates = {"EUR": 1.0, "USD": fx_day["USD"], "GBP": fx_day["GBP"]}
    amount_in_eur = float(amount) / rates[from_curr]
    return amount_in_eur * rates[to_curr]


# ─────────────────────────────────────────────────────────────────────────────
# 3. CARGA DE DATOS (sin cambios)
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(ttl=300)  # refresca desde Drive cada 5 minutos
def load_data():
    xl = pd.ExcelFile(load_excel_from_drive("datos"))

    df_char = pd.read_excel(xl, sheet_name="Characteristics")
    df_char.columns = df_char.columns.str.strip()
    df_char['Fecha Commitment'] = pd.to_datetime(df_char['Fecha Commitment'], dayfirst=True)
    df_char['Vintage']    = pd.to_numeric(df_char['Vintage'], errors='coerce')
    df_char['Commitment'] = pd.to_numeric(df_char['Commitment'], errors='coerce')
    df_char['Currency']   = df_char['Currency'].astype(str).str.strip().str.upper()
    df_char['GP']         = df_char['GP'].astype(str).str.strip()
    df_char['Fund']       = df_char['Fund'].astype(str).str.strip()
    df_char['Strategy']   = df_char['Strategy'].astype(str).str.strip()
    df_char['Geography']  = df_char['Geography'].astype(str).str.strip() if 'Geography' in df_char.columns else 'Unknown'
    # Fecha 1er Call en Characteristics (fondos) — columna opcional
    col_cc = None
    for c in df_char.columns:
        c_clean = c.strip().lower().replace('°','').replace('º','').replace('  ',' ')
        if 'fecha' in c_clean and ('1' in c_clean or 'primer' in c_clean) and 'call' in c_clean:
            col_cc = c
            break
    if col_cc:
        df_char['Fecha 1er Call'] = pd.to_datetime(df_char[col_cc], dayfirst=True, errors='coerce')
    else:
        df_char['Fecha 1er Call'] = pd.NaT
    df_char['Fecha 1er Call'] = df_char['Fecha 1er Call'].fillna(df_char['Fecha Commitment'])
    df_char = df_char.dropna(subset=['Commitment', 'Vintage'])

    # ── Characteristics_CoInv (pestaña separada para co-inversiones) ──────────
    df_coinv = pd.DataFrame()
    if 'Characteristics_CoInv' in xl.sheet_names:
        df_coinv = pd.read_excel(xl, sheet_name="Characteristics_CoInv")
        df_coinv.columns = df_coinv.columns.str.strip()
        # Columnas base
        for col_dt in ['Fecha Commitment']:
            if col_dt in df_coinv.columns:
                df_coinv[col_dt] = pd.to_datetime(df_coinv[col_dt], dayfirst=True, errors='coerce')
        for col_num in ['Commitment', 'Vintage', 'UW TVPI', 'New TVPI']:
            if col_num in df_coinv.columns:
                df_coinv[col_num] = pd.to_numeric(df_coinv[col_num], errors='coerce')
        for col_dt2 in ['Fecha 1° Capital Call', 'UW Exit Date', 'New Exit Date']:
            if col_dt2 in df_coinv.columns:
                df_coinv[col_dt2] = pd.to_datetime(df_coinv[col_dt2], dayfirst=True, errors='coerce')
        for col_str in ['Fund', 'GP', 'Strategy', 'Currency', 'Geography']:
            if col_str in df_coinv.columns:
                df_coinv[col_str] = df_coinv[col_str].astype(str).str.strip()
        if 'Currency' in df_coinv.columns:
            df_coinv['Currency'] = df_coinv['Currency'].str.upper()
        if 'Strategy' not in df_coinv.columns:
            df_coinv['Strategy'] = 'Single Co-Inv'
        df_coinv = df_coinv.dropna(subset=['Commitment'])

        # ── CRÍTICO: agregar co-inversiones a df_char para que TODAS las
        # pestañas existentes (Status, Portfolio, IRR, etc.) sigan funcionando ──
        df_coinv_base = df_coinv.copy()

        # Fecha 1er Call
        if 'Fecha 1° Capital Call' in df_coinv_base.columns:
            df_coinv_base['Fecha 1er Call'] = df_coinv_base['Fecha 1° Capital Call']
        elif 'Fecha Commitment' in df_coinv_base.columns:
            df_coinv_base['Fecha 1er Call'] = df_coinv_base['Fecha Commitment']
        else:
            df_coinv_base['Fecha 1er Call'] = pd.NaT
        if 'Fecha Commitment' in df_coinv_base.columns:
            df_coinv_base['Fecha 1er Call'] = df_coinv_base['Fecha 1er Call'].fillna(
                df_coinv_base['Fecha Commitment'])

        # Vintage obligatorio para df_char — usar año de Fecha Commitment si falta
        if 'Vintage' not in df_coinv_base.columns or df_coinv_base['Vintage'].isna().all():
            if 'Fecha Commitment' in df_coinv_base.columns:
                df_coinv_base['Vintage'] = pd.to_datetime(
                    df_coinv_base['Fecha Commitment'], errors='coerce').dt.year
        else:
            df_coinv_base['Vintage'] = pd.to_numeric(df_coinv_base['Vintage'], errors='coerce')
            if 'Fecha Commitment' in df_coinv_base.columns:
                df_coinv_base['Vintage'] = df_coinv_base['Vintage'].fillna(
                    pd.to_datetime(df_coinv_base['Fecha Commitment'], errors='coerce').dt.year)

        if 'Geography' not in df_coinv_base.columns:
            df_coinv_base['Geography'] = 'Unknown'

        # Columnas mínimas para df_char
        cols_base = ['Fund','GP','Strategy','Currency','Vintage','Commitment',
                     'Fecha Commitment','Geography','Fecha 1er Call']
        coinv_for_char = df_coinv_base[[c for c in cols_base
                                        if c in df_coinv_base.columns]].copy()
        coinv_for_char = coinv_for_char.dropna(subset=['Commitment'])

        # Unir al df_char principal — evitar duplicados
        existing_funds = set(df_char['Fund'].tolist())
        coinv_new = coinv_for_char[~coinv_for_char['Fund'].isin(existing_funds)]
        if not coinv_new.empty:
            df_char = pd.concat([df_char, coinv_new], ignore_index=True)

    df_flows_raw = pd.read_excel(xl, sheet_name="Cashflows")
    df_flows_raw.columns = df_flows_raw.columns.str.strip()
    df_flows_raw['Date']   = pd.to_datetime(df_flows_raw['Date'], dayfirst=True)
    df_flows_raw['Amount'] = pd.to_numeric(df_flows_raw['Amount'], errors='coerce').fillna(0)
    df_flows_raw['Fund']   = df_flows_raw['Fund'].astype(str).str.strip()
    df_flows_raw['Type']   = df_flows_raw['Type'].astype(str).str.strip()
    df_flows_raw = df_flows_raw.dropna(subset=['Date'])

    # ── Validación de calidad de datos ───────────────────────────────────────
    # Capital Calls positivos (deberían ser negativos)
    bad_calls = df_flows_raw[
        df_flows_raw['Type'].str.contains('Call', case=False) &
        (df_flows_raw['Amount'] > 0)
    ][['Fund','Date','Type','Amount']].copy()

    # Distribuciones negativas (deberían ser positivas)
    bad_dists = df_flows_raw[
        df_flows_raw['Type'].str.contains('Dist', case=False) &
        (df_flows_raw['Amount'] < 0)
    ][['Fund','Date','Type','Amount']].copy()

    # Guardar en session_state para mostrar en la UI
    import streamlit as _st
    if not bad_calls.empty or not bad_dists.empty:
        _st.session_state['_data_quality_errors'] = {
            'bad_calls': bad_calls,
            'bad_dists': bad_dists,
        }
    else:
        _st.session_state['_data_quality_errors'] = None

    return df_char, df_flows_raw, df_coinv


# ─────────────────────────────────────────────────────────────────────────────
# 4. COMPUTE PORTFOLIO (sin cambios)
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(show_spinner="Calculando portfolio...")
def compute_portfolio(_df_char_filt, _df_flows_raw, report_curr, as_of_date_dt, _fx_map,
                      cache_key=None):  # cache_key sin _ para que Streamlit lo use como hash
    stats_list = []
    all_cf_list = []
    fx_today = _fx_map.get(as_of_date_dt.date(), {"USD": 1.08, "GBP": 0.86})

    for _, f_meta in _df_char_filt.iterrows():
        fund, f_curr = f_meta['Fund'], f_meta['Currency']
        f_flows = _df_flows_raw[
            (_df_flows_raw['Fund'] == fund) & (_df_flows_raw['Date'] <= as_of_date_dt)
        ].copy()

        # Si no hay flujos hasta la fecha de corte, saltar
        if f_flows.empty:
            continue

        f_flows['Amt_Rep'] = [
            convert_amount(r['Amount'], f_curr, report_curr, _fx_map.get(r['Date'].date(), fx_today))
            for _, r in f_flows.iterrows()
        ]
        calls_rep = abs(sum(f_flows[f_flows['Type'].str.contains('Call', case=False)]['Amt_Rep']))
        dists_rep = sum(f_flows[f_flows['Type'].str.contains('Dist', case=False)]['Amt_Rep'])
        nav_entries = f_flows[f_flows['Type'].str.contains('NAV', case=False)].sort_values('Date')
        if not nav_entries.empty:
            ln = nav_entries.iloc[-1]
            later_f = f_flows[f_flows['Date'] > ln['Date']]
            nav_loc = (float(ln['Amount'])
                       + abs(float(later_f[later_f['Type'].str.contains('Call', case=False)]['Amount'].sum()))
                       - float(later_f[later_f['Type'].str.contains('Dist', case=False)]['Amount'].sum()))
            nav_rep = convert_amount(nav_loc, f_curr, report_curr, fx_today)
            nav_period = f"{(ln['Date'].month - 1) // 3 + 1}Q{ln['Date'].strftime('%y')}"
        else:
            nav_rep, nav_period = calls_rep, "Cost"

        # Commitment solo si Fecha Commitment <= as_of_date_dt
        if f_meta['Fecha Commitment'] <= as_of_date_dt:
            fx_commit = _fx_map.get(f_meta['Fecha Commitment'].date(), fx_today)
            comm_rep = convert_amount(f_meta['Commitment'], f_curr, report_curr, fx_commit)
        else:
            comm_rep = 0  # No comprometido aún en la fecha de corte

        cf_df = f_flows[~f_flows['Type'].str.contains('NAV', case=False)][['Date', 'Amt_Rep']].copy()
        all_cf_list.append(cf_df)
        try:
            nav_date = pd.Timestamp(as_of_date_dt)
            cf_agg = cf_df.groupby('Date')['Amt_Rep'].sum().reset_index()
            if nav_date in cf_agg['Date'].values:
                cf_agg.loc[cf_agg['Date'] == nav_date, 'Amt_Rep'] += float(nav_rep)
                tir_rep = xirr(cf_agg['Date'].tolist(), cf_agg['Amt_Rep'].tolist()) * 100
            else:
                tir_rep = xirr(cf_agg['Date'].tolist() + [nav_date],
                               cf_agg['Amt_Rep'].tolist() + [float(nav_rep)]) * 100
        except:
            tir_rep = 0
        stats_list.append({
            "Fund": fund, "GP": f_meta['GP'], "Strategy": f_meta['Strategy'],
            "Vintage": f_meta['Vintage'], "Periodo NAV": nav_period,
            "Commitment": comm_rep, "Paid-In": calls_rep, "Unfunded": max(comm_rep - calls_rep, 0),
            "Distributed": dists_rep, "NAV": nav_rep, "Total Value": dists_rep + nav_rep,
            "IRR %": tir_rep,
            "TVPI": (dists_rep + nav_rep) / (calls_rep if calls_rep > 0 else 1),
            "DPI": dists_rep / (calls_rep if calls_rep > 0 else 1),
            "committed": f_meta['Fecha Commitment'] <= as_of_date_dt,  # flag para contar inversiones
        })
    return pd.DataFrame(stats_list), all_cf_list



@st.cache_data(show_spinner=False)
def _calc_quarterly_evolutions(calc_curr, _df_final, _df_flows_raw, _df_char,
                                as_of_date_dt, _fx_map, fx_today, _cache_key, _version="v4"):
    """Cálculo cacheado de evoluciones trimestrales IRR/TVPI/DPI/Rent."""
    from pyxirr import xirr as _xirr
    irr_ev_l, tvpi_ev_l, dpi_ev_l, ret_ev_l = [], [], [], []
    if _df_final.empty:
        return irr_ev_l, tvpi_ev_l, dpi_ev_l, ret_ev_l
    nav_flows = _df_flows_raw[_df_flows_raw['Type'].str.contains('NAV', case=False)]
    last_nav_date = nav_flows['Date'].max() if not nav_flows.empty else as_of_date_dt
    q_dates = pd.date_range(start=_df_flows_raw['Date'].min(), end=last_nav_date, freq='QE')

    for f_name in _df_final['Fund'].tolist():
        f_i, f_t, f_d, f_r = {'Fund': f_name}, {'Fund': f_name}, {'Fund': f_name}, {'Fund': f_name}
        f_fl_raw = _df_flows_raw[_df_flows_raw['Fund'] == f_name].copy()
        f_mt_rows = _df_char[_df_char['Fund'] == f_name]
        if f_mt_rows.empty: continue
        f_mt = f_mt_rows.iloc[0]
        f_curr_ind, p_n_rep = f_mt['Currency'], 0
        # Usar fecha del primer flujo real en vez de Fecha Commitment
        # (puede haber fondos con flows antes de la firma oficial)
        first_flow_date = f_fl_raw['Date'].min() if not f_fl_raw.empty else f_mt['Fecha Commitment']

        def cv(amt, fx_day, f_curr_ind=f_curr_ind):
            if calc_curr == "Local":
                return float(amt)
            return convert_amount(amt, f_curr_ind, calc_curr, fx_day)

        for q_d in q_dates:
            if first_flow_date > q_d:
                for d in [f_i, f_t, f_d, f_r]:
                    d[q_d.strftime('%d-%m-%Y')] = None
                continue
            q_fx = _fx_map.get(q_d.date(), fx_today)
            q_f  = f_fl_raw[f_fl_raw['Date'] <= q_d].copy()
            q_f['Amt_R'] = [
                cv(r['Amount'], _fx_map.get(r['Date'].date(), fx_today))
                for _, r in q_f.iterrows()
            ]
            q_nv_l = q_f[q_f['Type'].str.contains('NAV', case=False)].sort_values('Date')
            if not q_nv_l.empty:
                qn  = q_nv_l.iloc[-1]
                n_l = (float(qn['Amount'])
                       + abs(float(q_f[q_f['Date'] > qn['Date']][q_f['Type'].str.contains('Call', case=False)]['Amount'].sum()))
                       - float(q_f[q_f['Date'] > qn['Date']][q_f['Type'].str.contains('Dist', case=False)]['Amount'].sum()))
                n_r = cv(n_l, q_fx)
            else:
                n_r = abs(sum(q_f[q_f['Type'].str.contains('Call', case=False)]['Amt_R']))
            # Verificar si hay un NAV en este trimestre (no necesariamente en el último día)
            qs_check = q_d - pd.tseries.offsets.QuarterEnd()
            nav_this_q = q_f[
                q_f['Type'].str.contains('NAV', case=False) &
                (q_f['Date'] > qs_check) & (q_f['Date'] <= q_d)
            ]
            if not nav_this_q.empty:
                qs     = q_d - pd.tseries.offsets.QuarterEnd()
                cf_q   = f_fl_raw[(f_fl_raw['Date'] > qs) & (f_fl_raw['Date'] <= q_d)].copy()
                qc     = cv(abs(float(cf_q[cf_q['Type'].str.contains('Call', case=False)]['Amount'].sum())), q_fx)
                qd_val = cv(float(cf_q[cf_q['Type'].str.contains('Dist', case=False)]['Amount'].sum()), q_fx)
                if p_n_rep == 0:
                    # Primer trimestre activo: denominador = capital neto invertido (calls - dists)
                    denom = qc - qd_val
                else:
                    denom = p_n_rep - qd_val + qc
                if denom <= 0:
                    f_r[q_d.strftime('%d-%m-%Y')] = None
                else:
                    ret = ((n_r / denom) - 1) * 100
                    f_r[q_d.strftime('%d-%m-%Y')] = ret if -100 < ret < 500 else None
            else:
                f_r[q_d.strftime('%d-%m-%Y')] = None
            p_n_rep = n_r
            tqc = abs(sum(q_f[q_f['Type'].str.contains('Call', case=False)]['Amt_R']))
            tqd = sum(q_f[q_f['Type'].str.contains('Dist', case=False)]['Amt_R'])
            f_t[q_d.strftime('%d-%m-%Y')] = (tqd + n_r) / (tqc if tqc > 0 else 1)
            f_d[q_d.strftime('%d-%m-%Y')] = tqd / (tqc if tqc > 0 else 1)
            try:
                f_i[q_d.strftime('%d-%m-%Y')] = _xirr(
                    q_f[~q_f['Type'].str.contains('NAV', case=False)]['Date'].tolist() + [q_d],
                    q_f[~q_f['Type'].str.contains('NAV', case=False)]['Amt_R'].tolist() + [float(n_r)]
                ) * 100
            except:
                f_i[q_d.strftime('%d-%m-%Y')] = 0
        irr_ev_l.append(f_i); tvpi_ev_l.append(f_t)
        dpi_ev_l.append(f_d); ret_ev_l.append(f_r)
    return irr_ev_l, tvpi_ev_l, dpi_ev_l, ret_ev_l




@st.cache_data(show_spinner=False)
def _calc_pooled_irr(fund_list, nav_total, report_curr, as_of_date_dt,
                     _df_flows_raw, _df_char, _fx_map, fx_today):
    flows = []
    for fund in fund_list:
        f_meta_rows = _df_char[_df_char['Fund'] == fund]
        if f_meta_rows.empty: continue
        f_meta = f_meta_rows.iloc[0]
        f_curr = f_meta['Currency']
        f_flows = _df_flows_raw[
            (_df_flows_raw['Fund'] == fund) & (_df_flows_raw['Date'] <= as_of_date_dt)
        ].copy()
        f_flows['Amt_Rep'] = [
            convert_amount(r['Amount'], f_curr, report_curr, _fx_map.get(r['Date'].date(), fx_today))
            for _, r in f_flows.iterrows()
        ]
        cf = f_flows[~f_flows['Type'].str.contains('NAV', case=False)][['Date', 'Amt_Rep']].copy()
        flows.append(cf)
    if not flows:
        return 0
    agg = pd.concat(flows, ignore_index=True).groupby('Date')['Amt_Rep'].sum().reset_index()
    try:
        nav_date = pd.Timestamp(as_of_date_dt)
        if nav_date in agg['Date'].values:
            agg.loc[agg['Date'] == nav_date, 'Amt_Rep'] += nav_total
            irr_dates, irr_amounts = agg['Date'].tolist(), agg['Amt_Rep'].tolist()
        else:
            irr_dates   = agg['Date'].tolist() + [nav_date]
            irr_amounts = agg['Amt_Rep'].tolist() + [nav_total]
        irr = xirr(irr_dates, irr_amounts) * 100
        return irr if irr > -99 else 0
    except:
        return 0


try:
    # ─────────────────────────────────────────────────────────────────────────
    # 5. SIDEBAR — diseño CobaltLP
    # ─────────────────────────────────────────────────────────────────────────
    df_char, df_flows_raw, df_coinv = load_data()

    # ── Alarma de calidad de datos ────────────────────────────────────────────
    dq = st.session_state.get('_data_quality_errors')
    if dq:
        bad_calls = dq['bad_calls']
        bad_dists = dq['bad_dists']
        with st.expander("⚠️ Errores de datos detectados en Cashflows — haz clic para ver", expanded=False):
            if not bad_calls.empty:
                st.markdown(
                    f"**🔴 Capital Calls con signo positivo** — deberían ser negativos "
                    f"({len(bad_calls)} flujo{'s' if len(bad_calls)>1 else ''}):"
                )
                bad_calls_show = bad_calls.copy()
                bad_calls_show['Date'] = bad_calls_show['Date'].dt.strftime('%d/%m/%Y')
                st.dataframe(
                    bad_calls_show.style.format({'Amount': '{:,.0f}'}),
                    use_container_width=True,
                    height=min(150, 40 + len(bad_calls_show) * 35),
                )
            if not bad_dists.empty:
                st.markdown(
                    f"**🔴 Distribuciones con signo negativo** — deberían ser positivas "
                    f"({len(bad_dists)} flujo{'s' if len(bad_dists)>1 else ''}):"
                )
                bad_dists_show = bad_dists.copy()
                bad_dists_show['Date'] = bad_dists_show['Date'].dt.strftime('%d/%m/%Y')
                st.dataframe(
                    bad_dists_show.style.format({'Amount': '{:,.0f}'}),
                    use_container_width=True,
                    height=min(150, 40 + len(bad_dists_show) * 35),
                )
            st.caption(
                "Corrige estos valores en la hoja **Cashflows** de `datos.xlsx`. "
                "Los signos incorrectos distorsionan el cálculo de TIR y los totales."
            )

    # Logo header
    st.sidebar.markdown("""
    <div class="sb-logo">
      <div class="sb-logo-mark">FO</div>
      <div>
        <div class="sb-logo-name">Family Office OS</div>
        <div class="sb-logo-sub">Alternative Assets</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Configuración (moneda + fecha) ────────────────────────────────────────
    st.sidebar.markdown('<div class="sb-section-label">Configuración</div>',
                        unsafe_allow_html=True)

    report_curr = st.sidebar.selectbox("Moneda de Reporte", ["USD", "EUR", "GBP"],
                                        label_visibility="visible")
    curr_sym    = {"USD": "$", "EUR": "€", "GBP": "£"}[report_curr]
    as_of_date  = st.sidebar.date_input("Fecha de Corte",
                                         value=df_flows_raw['Date'].max().date())
    as_of_date_dt = pd.to_datetime(as_of_date)

    # fx_map cubre desde el primer flujo hasta la fecha máxima de datos
    # (necesario para Cash Flows que muestra todos los trimestres independiente del corte)
    fx_map_end = max(pd.Timestamp(as_of_date), df_flows_raw['Date'].max())
    fx_map = get_fx_map_institutional(
        df_flows_raw['Date'].min() - timedelta(days=30), fx_map_end
    )
    # fx_map ahora siempre devuelve un mapa válido (con fallback si la API falla)
    # Verificar si se usaron tasas de fallback mostrando un aviso suave
    as_of_key = as_of_date.date() if hasattr(as_of_date, 'date') else as_of_date
    fx_today  = fx_map.get(as_of_key, {"USD": 1.08, "GBP": 0.86})
    if fx_today == {"USD": 1.08, "GBP": 0.86}:
        st.sidebar.caption("⚠️ Tasas BCE no disponibles — usando EUR/USD 1.08, EUR/GBP 0.86")

    # Tasas BCE pill
    st.sidebar.markdown(f"""
    <div class="sb-rates">
      <div class="sb-rates-title">Tasas BCE</div>
      <div class="sb-rate-row">
        <span class="sb-rate-key">EUR / USD</span>
        <span class="sb-rate-val">{fx_today['USD']:.4f}</span>
      </div>
      <div class="sb-rate-row">
        <span class="sb-rate-key">EUR / GBP</span>
        <span class="sb-rate-val">{fx_today['GBP']:.4f}</span>
      </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Filtros — estilo CobaltLP (expanders colapsables) ────────────────────
    st.sidebar.markdown('<div class="sb-section-label">Chart Filters</div>',
                        unsafe_allow_html=True)

    all_gps    = sorted(df_char["GP"].unique().tolist())
    all_funds  = sorted(df_char["Fund"].unique().tolist())
    all_strats = sorted(df_char["Strategy"].unique().tolist())
    all_geos   = sorted(df_char["Geography"].unique().tolist()) if "Geography" in df_char.columns else []

    # Inicializar selecciones en session_state
    if "gp_sel"    not in st.session_state: st.session_state["gp_sel"]    = all_gps
    if "fund_sel"  not in st.session_state: st.session_state["fund_sel"]  = all_funds
    if "strat_sel" not in st.session_state: st.session_state["strat_sel"] = all_strats
    if "geo_sel"   not in st.session_state: st.session_state["geo_sel"]   = all_geos

    # Expander GP
    with st.sidebar.expander("⬥  GP / Manager", expanded=False):
        sel_all_gp = st.checkbox("Seleccionar todos", value=True, key="chk_gp")
        gp_filt = st.multiselect(
            "GP", all_gps,
            default=all_gps if sel_all_gp else st.session_state["gp_sel"],
            key="ms_gp", label_visibility="collapsed"
        )

    # Expander Estrategia
    with st.sidebar.expander("⬥  Style / Focus", expanded=False):
        sel_all_strat = st.checkbox("Seleccionar todas", value=True, key="chk_strat")
        strat_filt = st.multiselect(
            "Estrategia", all_strats,
            default=all_strats if sel_all_strat else st.session_state["strat_sel"],
            key="ms_strat", label_visibility="collapsed"
        )

    # Expander Vintage Year
    all_vintages = sorted(df_char["Vintage"].dropna().unique().astype(int).tolist())
    with st.sidebar.expander("⬥  Vintage Year", expanded=False):
        sel_all_vint = st.checkbox("Seleccionar todos", value=True, key="chk_vint")
        vint_filt = st.multiselect(
            "Vintage", all_vintages,
            default=all_vintages if sel_all_vint else [],
            key="ms_vint", label_visibility="collapsed"
        )

    # Expander Fondo
    with st.sidebar.expander("⬥  Fund", expanded=False):
        sel_all_fund = st.checkbox("Seleccionar todos", value=True, key="chk_fund")
        fund_filt = st.multiselect(
            "Fondo", all_funds,
            default=all_funds if sel_all_fund else st.session_state["fund_sel"],
            key="ms_fund", label_visibility="collapsed"
        )

    # Expander Geografía (si existe la columna)
    if all_geos:
        with st.sidebar.expander("⬥  Geography", expanded=False):
            sel_all_geo = st.checkbox("Seleccionar todas", value=True, key="chk_geo")
            geo_filt = st.multiselect(
                "Geografía", all_geos,
                default=all_geos if sel_all_geo else st.session_state["geo_sel"],
                key="ms_geo", label_visibility="collapsed"
            )
    else:
        geo_filt = []

    # Botón Clear All (filtros)
    st.sidebar.markdown('<div class="sb-clear-btn">', unsafe_allow_html=True)
    if st.sidebar.button("Clear All", key="btn_clear"):
        for key in ["chk_gp", "chk_fund", "chk_strat", "chk_vint", "chk_geo"]:
            if key in st.session_state:
                st.session_state[key] = False
        st.rerun()
    # Botón limpiar caché de cálculos
    if st.sidebar.button("🔄 Recalcular todo", key="btn_clear_cache",
                          help="Limpia el caché y recalcula todos los datos desde cero"):
        st.cache_data.clear()
        st.rerun()
    if st.sidebar.button("☁️ Recargar datos Drive", key="btn_reload_drive",
                          help="Fuerza releer datos.xlsx desde Google Drive"):
        load_data.clear()
        st.rerun()
    st.sidebar.markdown('</div>', unsafe_allow_html=True)

    # ─────────────────────────────────────────────────────────────────────────
    # 6. APLICAR FILTROS
    # ─────────────────────────────────────────────────────────────────────────
    # Para cada filtro: si el multiselect está vacío Y el checkbox "todos" está
    # desmarcado → filtrar por lista vacía (mostrar nada).
    # Si el multiselect está vacío Y el checkbox "todos" está marcado → no filtrar.

    mask_gp    = df_char["GP"].isin(gp_filt)       if gp_filt    else (pd.Series(True,  index=df_char.index) if sel_all_gp   else pd.Series(False, index=df_char.index))
    mask_fund  = df_char["Fund"].isin(fund_filt)    if fund_filt  else (pd.Series(True,  index=df_char.index) if sel_all_fund else pd.Series(False, index=df_char.index))
    mask_strat = df_char["Strategy"].isin(strat_filt) if strat_filt else (pd.Series(True, index=df_char.index) if sel_all_strat else pd.Series(False, index=df_char.index))

    df_char_filt = df_char[mask_gp & mask_fund & mask_strat].copy()

    # Filtro vintage — convertir a int para evitar mismatch float vs int
    if vint_filt:
        df_char_filt = df_char_filt[
            df_char_filt["Vintage"].fillna(-1).astype(int).isin(vint_filt)
        ]
    elif not sel_all_vint:
        df_char_filt = df_char_filt.iloc[0:0]

    # Filtro geografía
    if all_geos:
        if geo_filt:
            df_char_filt = df_char_filt[df_char_filt["Geography"].isin(geo_filt)]
        elif not sel_all_geo:
            df_char_filt = df_char_filt.iloc[0:0]

    data_hash = hash(tuple(df_flows_raw[['Fund','Date','Type','Amount']].values.tobytes()))
    df_final, all_flows_rep_global = compute_portfolio(
        df_char_filt.reset_index(drop=True),
        df_flows_raw,
        report_curr,
        as_of_date_dt,
        fx_map,
        (tuple(sorted(df_char_filt['Fund'].tolist())),
         str(as_of_date_dt.date()), report_curr,
         round(fx_today.get('USD', 1.0), 6),
         data_hash),
    )
    if not df_final.empty:
        df_final.index = range(1, len(df_final) + 1)

    # ─────────────────────────────────────────────────────────────────────────
    # 7. FUNCIONES DE ANÁLISIS (sin cambios)
    # ─────────────────────────────────────────────────────────────────────────
    def calc_quarterly_evolutions(calc_curr):
        # Hash del contenido de los datos para detectar cambios en el Excel
        data_hash = hash(tuple(df_flows_raw[['Fund','Date','Type','Amount']].values.tobytes()))
        return _calc_quarterly_evolutions(
            calc_curr, df_final, df_flows_raw, df_char,
            as_of_date_dt, fx_map, fx_today,
            (tuple(sorted(df_final['Fund'].tolist())),
             str(as_of_date_dt.date()), calc_curr,
             round(fx_today.get('USD', 1.0), 6),
             data_hash),
            "v5",
        )

    def calc_pooled_irr(group_df):
        fund_list  = tuple(sorted(group_df['Fund'].tolist()))
        nav_total  = float(group_df['NAV'].sum())
        return _calc_pooled_irr(
            fund_list, nav_total, report_curr, as_of_date_dt,
            df_flows_raw, df_char, fx_map, fx_today
        )

    v_stats = df_final.groupby('Vintage').agg({
        'Commitment': 'sum', 'Paid-In': 'sum', 'Unfunded': 'sum',
        'Distributed': 'sum', 'NAV': 'sum', 'Total Value': 'sum'
    }).reset_index()
    v_stats['TVPI']  = v_stats['Total Value'] / v_stats['Paid-In']
    v_stats['DPI']   = v_stats['Distributed'] / v_stats['Paid-In']
    v_stats['IRR %'] = v_stats['Vintage'].apply(
        lambda v: calc_pooled_irr(df_final[df_final['Vintage'] == v])
    )

    s_stats = df_final.groupby('Strategy').agg({
        'Commitment': 'sum', 'Paid-In': 'sum', 'Unfunded': 'sum',
        'Distributed': 'sum', 'NAV': 'sum', 'Total Value': 'sum'
    }).reset_index()
    s_stats['TVPI']  = s_stats['Total Value'] / s_stats['Paid-In']
    s_stats['DPI']   = s_stats['Distributed'] / s_stats['Paid-In']
    s_stats['IRR %'] = s_stats['Strategy'].apply(
        lambda s: calc_pooled_irr(df_final[df_final['Strategy'] == s])
    )

    def calc_wav_duration(group_df):
        funds = group_df['Fund'].tolist()
        flows_list = []
        for fund in funds:
            f_meta  = df_char[df_char['Fund'] == fund].iloc[0]
            f_curr  = f_meta['Currency']
            f_flows = df_flows_raw[
                (df_flows_raw['Fund'] == fund) & (df_flows_raw['Date'] <= as_of_date_dt)
                & (df_flows_raw['Type'].str.contains('Call', case=False))
            ].copy()
            if f_flows.empty:
                continue
            f_flows['Amt_Rep'] = [
                abs(convert_amount(r['Amount'], f_curr, report_curr, fx_map.get(r['Date'].date(), fx_today)))
                for _, r in f_flows.iterrows()
            ]
            f_flows['Years'] = (as_of_date_dt - f_flows['Date']).dt.days / 365.25
            flows_list.append(f_flows[['Amt_Rep', 'Years']])
        if not flows_list:
            return 0
        agg     = pd.concat(flows_list, ignore_index=True)
        total_w = agg['Amt_Rep'].sum()
        if total_w == 0:
            return 0
        return (agg['Amt_Rep'] * agg['Years']).sum() / total_w

    s_stats['Duration (yrs)'] = s_stats['Strategy'].apply(
        lambda s: calc_wav_duration(df_final[df_final['Strategy'] == s])
    )
    _dur_total = calc_wav_duration(df_final)

    # ─────────────────────────────────────────────────────────────────────────
    # 8. PERFORMANCE GLOBAL POOLED (sin cambios)
    # ─────────────────────────────────────────────────────────────────────────
    t_comm = df_final[df_final['committed'] == True]['Commitment'].sum()
    t_paid = df_final['Paid-In'].sum()
    t_dist = df_final['Distributed'].sum()
    t_nav  = df_final['NAV'].sum()
    try:
        f_p_agg = pd.concat(all_flows_rep_global, ignore_index=True).groupby('Date')['Amt_Rep'].sum().reset_index()
        nav_date = pd.Timestamp(as_of_date_dt)
        if nav_date in f_p_agg['Date'].values:
            f_p_agg.loc[f_p_agg['Date'] == nav_date, 'Amt_Rep'] += float(t_nav)
            irr_dates   = f_p_agg['Date'].tolist()
            irr_amounts = f_p_agg['Amt_Rep'].tolist()
        else:
            irr_dates   = f_p_agg['Date'].tolist() + [nav_date]
            irr_amounts = f_p_agg['Amt_Rep'].tolist() + [float(t_nav)]
        g_irr = xirr(irr_dates, irr_amounts) * 100
        if g_irr < -99:
            g_irr = 0
    except:
        g_irr = 0

    # ─────────────────────────────────────────────────────────────────────────
    # 9. TABS — contenido 100% original sin cambios
    # ─────────────────────────────────────────────────────────────────────────
    TAB_NAMES = ["📊 Status", "📅 Vintage", "🎯 Estrategia", "💼 Portfolio",
                 "📉 IRR", "📈 TVPI", "💰 DPI", "🔄 Rent.",
                 "💸 Cash Flows", "📆 Commitment Pace",
                 "📍 Point in Time", "🔮 Simulación"]

    # Inicializar pestaña activa en session_state
    if 'active_tab' not in st.session_state:
        st.session_state['active_tab'] = TAB_NAMES[0]

    # CSS para el radio horizontal que simula tabs
    st.markdown("""
    <style>
    div[data-testid="stHorizontalBlock"] > div[data-testid="column"] {
        padding: 0 !important;
    }
    div.tab-radio > div[role="radiogroup"] {
        display: flex; flex-wrap: wrap; gap: 2px;
        border-bottom: 2px solid #e0e6f0;
        padding-bottom: 0; margin-bottom: 16px;
    }
    div.tab-radio > div[role="radiogroup"] > label {
        padding: 8px 14px !important;
        border-radius: 6px 6px 0 0 !important;
        font-size: 13px !important;
        font-weight: 500 !important;
        color: #5a7fa0 !important;
        border: 1px solid transparent !important;
        border-bottom: none !important;
        cursor: pointer;
        margin-bottom: -2px !important;
        background: transparent !important;
        transition: all 0.15s;
    }
    div.tab-radio > div[role="radiogroup"] > label:hover {
        color: #002060 !important;
        background: #f0f4fa !important;
    }
    div.tab-radio > div[role="radiogroup"] > label[data-baseweb="radio"]:has(input:checked),
    div.tab-radio > div[role="radiogroup"] > label:has(input:checked) {
        color: #1a5fd4 !important;
        border-color: #e0e6f0 !important;
        border-bottom-color: white !important;
        background: white !important;
        font-weight: 600 !important;
    }
    div.tab-radio > div[role="radiogroup"] > label > div:first-child {
        display: none !important;
    }
    </style>
    """, unsafe_allow_html=True)

    st.markdown('<div class="tab-radio">', unsafe_allow_html=True)
    active_tab = st.radio(
        "navegación", TAB_NAMES,
        index=TAB_NAMES.index(st.session_state['active_tab']),
        horizontal=True,
        label_visibility="collapsed",
        key="tab_radio"
    )
    st.markdown('</div>', unsafe_allow_html=True)
    st.session_state['active_tab'] = active_tab

    # Contenedor de la pestaña activa
    def tab_active(name):
        return active_tab == name

    if tab_active("📊 Status"):
        t_tv   = t_dist + t_nav
        t_unf  = max(t_comm - t_paid, 0)
        t_gl   = t_tv - t_paid
        g_tvpi = t_tv / t_paid if t_paid > 0 else 0
        g_dpi  = t_dist / t_paid if t_paid > 0 else 0
        # N° inversiones y relaciones: solo fondos con Fecha Commitment <= as_of
        df_committed = df_final[df_final['committed'] == True]
        n_funds = len(df_committed)
        n_gps   = df_char_filt[df_char_filt['Fecha Commitment'] <= as_of_date_dt]['GP'].nunique()

        st.markdown("""
        <style>
        .kpi-row { display:flex; gap:0px; border:1.5px solid #d0d8e8; border-radius:10px; overflow:hidden; margin-bottom:22px; }
        .kpi-cell { flex:1; text-align:center; padding:14px 8px 12px 8px; border-right:1px solid #d0d8e8; background:#fff; }
        .kpi-cell:last-child { border-right:none; }
        .kpi-cell:hover { background:#f4f7fc; }
        .kpi-v { font-size:20px; font-weight:700; color:#002060; line-height:1.2; }
        .kpi-l { font-size:10px; font-weight:600; letter-spacing:0.8px; text-transform:uppercase; color:#7a93b8; margin-top:4px; text-decoration:underline; }
        .perf-box { border:1.5px solid #d0d8e8; border-radius:10px; padding:20px 16px; background:#fff; height:100%; }
        .perf-row { display:flex; justify-content:space-between; align-items:baseline; padding:12px 0; border-bottom:1px solid #eef1f7; }
        .perf-row:last-child { border-bottom:none; }
        .perf-metric { font-size:28px; font-weight:700; color:#002060; }
        .perf-label { font-size:11px; color:#7a93b8; font-weight:600; letter-spacing:0.8px; text-transform:uppercase; }
        </style>
        """, unsafe_allow_html=True)

        st.markdown(f"""
        <div class="kpi-row">
          <div class="kpi-cell"><div class="kpi-v">{curr_sym}{t_comm/1e6:,.1f} MM</div><div class="kpi-l">Commitment</div></div>
          <div class="kpi-cell"><div class="kpi-v">{n_funds}</div><div class="kpi-l">Inversiones</div></div>
          <div class="kpi-cell"><div class="kpi-v">{n_gps}</div><div class="kpi-l">Relaciones</div></div>
          <div class="kpi-cell"><div class="kpi-v">{curr_sym}{t_paid/1e6:,.1f} MM</div><div class="kpi-l">Paid-In</div></div>
          <div class="kpi-cell"><div class="kpi-v">{curr_sym}{t_unf/1e6:,.1f} MM</div><div class="kpi-l">Unfunded</div></div>
          <div class="kpi-cell"><div class="kpi-v">{curr_sym}{t_dist/1e6:,.1f} MM</div><div class="kpi-l">Distributed</div></div>
          <div class="kpi-cell"><div class="kpi-v">{curr_sym}{t_nav/1e6:,.1f} MM</div><div class="kpi-l">NAV</div></div>
          <div class="kpi-cell"><div class="kpi-v">{curr_sym}{t_tv/1e6:,.1f} MM</div><div class="kpi-l">Total Value</div></div>
          <div class="kpi-cell"><div class="kpi-v">{curr_sym}{t_gl/1e6:,.1f} MM</div><div class="kpi-l">Utilidad</div></div>
        </div>
        """, unsafe_allow_html=True)

        l_c, r_c = st.columns([4, 1])
        with l_c:
            fig = go.Figure()
            fig.add_trace(go.Bar(
                name='Commitment', x=['Commitment'], y=[t_comm],
                marker_color='#002060',
                text=[f"<b>{t_comm/1e6:,.1f}</b>"], textposition='inside',
                textfont=dict(color='white', size=13), width=0.6,
            ))
            fig.add_trace(go.Bar(
                name='Paid-In', x=['Allocation'], y=[t_paid],
                marker_color='#ED7D31',
                text=[f"<b>{t_paid/1e6:,.1f}</b>"], textposition='inside',
                textfont=dict(color='white', size=13), width=0.6,
            ))
            fig.add_trace(go.Bar(
                name='Unfunded', x=['Allocation'], y=[t_unf],
                marker_color='rgba(0,0,0,0)',
                marker_line_color='#002060', marker_line_width=2,
                text=[f"<b>{t_unf/1e6:,.1f}</b>"], textposition='outside',
                textfont=dict(color='#002060', size=13), width=0.6,
            ))
            for name, val, color in [
                ('Distributed', t_dist, '#92D050'),
                ('NAV',         t_nav,  '#FFC000'),
                ('Total Value', t_tv,   '#4472C4'),
            ]:
                fig.add_trace(go.Bar(
                    name=name, x=[name], y=[val],
                    marker_color=color,
                    text=[f"<b>{val/1e6:,.1f}</b>"], textposition='inside',
                    textfont=dict(color='white', size=13), width=0.6,
                ))
            max_val = max(t_comm, t_tv)
            fig.update_layout(
                barmode='stack', height=460,
                plot_bgcolor='white', paper_bgcolor='white',
                showlegend=True,
                legend=dict(orientation='h', y=1.08, font=dict(size=11)),
                title=dict(text=f"Portfolio Value Composition ({report_curr} MM)",
                           font=dict(size=14, color='#002060'), x=0),
                yaxis=dict(showgrid=True, gridcolor='#eef1f7', zeroline=False,
                           tickfont=dict(color='#999'), range=[0, max_val * 1.25]),
                xaxis=dict(tickfont=dict(size=13, color='#002060')),
                margin=dict(t=60, b=10, l=10, r=10),
            )
            st.plotly_chart(fig, use_container_width=True)

        with r_c:
            st.markdown(f"""
            <div class="perf-box" style="min-width:110px">
              <div class="perf-row"><div>
                <div class="perf-metric" style="font-size:22px">{g_irr:.2f}%</div>
                <div class="perf-label">Net IRR</div>
              </div></div>
              <div class="perf-row"><div>
                <div class="perf-metric" style="font-size:22px">{g_tvpi:.2f}x</div>
                <div class="perf-label">TVPI</div>
              </div></div>
              <div class="perf-row"><div>
                <div class="perf-metric" style="font-size:22px">{g_dpi:.2f}x</div>
                <div class="perf-label">DPI</div>
              </div></div>
              <div class="perf-row"><div>
                <div class="perf-metric" style="font-size:22px">{(t_nav/t_paid):.2f}x</div>
                <div class="perf-label">RVPI</div>
              </div></div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("---")

        div1, div2 = st.columns(2)
        ASSET_GROUPS = {
            "Private Equity": ["Buyout", "Secondaries", "Growth Equity", "Venture Capital", "Fund of Funds"],
            "Co-Investments": ["Single Co-Inv"],
            "Private Credit": ["Credit"],
            "Real Estate":    ["Real Estate"],
        }
        ASSET_COLORS = ['#002060', '#ED7D31', '#FFC000', '#4472C4']
        df_nav_asset = df_final.copy()
        def map_group(s):
            for g, strats in ASSET_GROUPS.items():
                if s in strats: return g
            return "Otros"
        df_nav_asset['Asset Group'] = df_nav_asset['Strategy'].apply(map_group)
        nav_by_asset = df_nav_asset.groupby('Asset Group')['NAV'].sum().reset_index()
        nav_by_asset = nav_by_asset[nav_by_asset['NAV'] > 0]

        with div1:
            fig_div1 = go.Figure(go.Pie(
                labels=nav_by_asset['Asset Group'], values=nav_by_asset['NAV'],
                marker=dict(colors=ASSET_COLORS[:len(nav_by_asset)]),
                textinfo='label+percent', textfont=dict(size=14),
                hovertemplate='%{label}<br>NAV: ' + curr_sym + '%{value:,.0f}<br>%{percent}<extra></extra>',
                hole=0.35,
            ))
            fig_div1.update_layout(
                title=dict(text="% NAV por Tipo de Activo", font=dict(size=14, color='#002060'), x=0),
                height=500, showlegend=True,
                legend=dict(orientation='h', y=-0.06, font=dict(size=12)),
                margin=dict(t=50, b=60, l=40, r=40),
            )
            st.plotly_chart(fig_div1, use_container_width=True)

        GEO_COLORS = ['#002060', '#ED7D31', '#4472C4', '#FFC000', '#92D050']
        df_nav_geo = df_final.merge(df_char_filt[['Fund','Geography']], on='Fund', how='left')
        df_nav_geo['Geography'] = df_nav_geo['Geography'].fillna('Unknown')
        nav_by_geo = df_nav_geo.groupby('Geography')['NAV'].sum().reset_index()
        nav_by_geo = nav_by_geo[nav_by_geo['NAV'] > 0]

        with div2:
            fig_div2 = go.Figure(go.Pie(
                labels=nav_by_geo['Geography'], values=nav_by_geo['NAV'],
                marker=dict(colors=GEO_COLORS[:len(nav_by_geo)]),
                textinfo='label+percent', textfont=dict(size=14),
                hovertemplate='%{label}<br>NAV: ' + curr_sym + '%{value:,.0f}<br>%{percent}<extra></extra>',
                hole=0.35,
            ))
            fig_div2.update_layout(
                title=dict(text="% NAV por Geografía", font=dict(size=14, color='#002060'), x=0),
                height=500, showlegend=True,
                legend=dict(orientation='h', y=-0.06, font=dict(size=12)),
                margin=dict(t=50, b=60, l=40, r=40),
            )
            st.plotly_chart(fig_div2, use_container_width=True)

    money_cols = ['Commitment', 'Paid-In', 'Unfunded', 'Distributed', 'NAV', 'Total Value']

    if tab_active("📅 Vintage"):
        fund_count_v = df_final.groupby('Vintage')['Fund'].count().reset_index().rename(columns={'Fund': '# Inv'})
        v_stats = v_stats.merge(fund_count_v, on='Vintage', how='left')
        v_stats = v_stats.sort_values('Vintage').reset_index(drop=True)

        total_comm_v = v_stats['Commitment'].sum()
        total_nav_v  = v_stats['NAV'].sum()
        total_paid_v = v_stats['Paid-In'].sum()
        total_funds_v = int(v_stats['# Inv'].sum())

        v_rows = []
        for _, row in v_stats.iterrows():
            paid = row['Paid-In']
            v_rows.append({
                'Vintage':          int(row['Vintage']),
                '# Inv':            int(row['# Inv']),
                'Commit (MM)':      row['Commitment'] / 1e6,
                'Paid In (MM)':     paid / 1e6,
                'Unfunded (MM)':    row['Unfunded'] / 1e6,
                'Distributed (MM)': row['Distributed'] / 1e6,
                'NAV (MM)':         row['NAV'] / 1e6,
                'IRR':              row['IRR %'],
                'TVPI':             row['TVPI'],
                'DPI':              row['DPI'],
                '% Comm':           row['Commitment'] / total_comm_v * 100 if total_comm_v > 0 else 0,
                '% NAV':            row['NAV'] / total_nav_v * 100 if total_nav_v > 0 else 0,
                '% Paid In':        paid / total_paid_v * 100 if total_paid_v > 0 else 0,
                'Duration (yrs)':   calc_wav_duration(df_final[df_final['Vintage'] == row['Vintage']]),
            })

        # Total row
        v_rows.append({
            'Vintage':          'Total',
            '# Inv':            total_funds_v,
            'Commit (MM)':      total_comm_v / 1e6,
            'Paid In (MM)':     total_paid_v / 1e6,
            'Unfunded (MM)':    v_stats['Unfunded'].sum() / 1e6,
            'Distributed (MM)': v_stats['Distributed'].sum() / 1e6,
            'NAV (MM)':         total_nav_v / 1e6,
            'IRR':              g_irr,
            'TVPI':             v_stats['Total Value'].sum() / total_paid_v if total_paid_v > 0 else 0,
            'DPI':              v_stats['Distributed'].sum() / total_paid_v if total_paid_v > 0 else 0,
            '% Comm':           100.0,
            '% NAV':            100.0,
            '% Paid In':        100.0,
            'Duration (yrs)':   _dur_total,
        })

        df_v = pd.DataFrame(v_rows)

        fmt_v = {
            '# Inv':            '{:.0f}',
            'Commit (MM)':      '{:,.1f}',
            'Paid In (MM)':     '{:,.1f}',
            'Unfunded (MM)':    '{:,.1f}',
            'Distributed (MM)': '{:,.1f}',
            'NAV (MM)':         '{:,.1f}',
            'IRR':              '{:.1f}%',
            'TVPI':             '{:.2f}x',
            'DPI':              '{:.2f}x',
            '% Comm':           '{:.1f}%',
            '% NAV':            '{:.1f}%',
            '% Paid In':        '{:.1f}%',
            'Duration (yrs)':   '{:.1f}',
        }

        def style_v(row):
            if str(row['Vintage']) == 'Total':
                return ['font-weight:bold; background-color:#eef1f7'] * len(row)
            return [''] * len(row)

        st.dataframe(
            df_v.style.format(fmt_v).apply(style_v, axis=1),
            use_container_width=True,
            height=min(50 + len(df_v) * 35, 600),
            hide_index=True,
        )
        excel_download_btn(df_v, "Vintage", f"vintage_{report_curr}.xlsx",
                           "Vintage", f"Portfolio por Vintage — {report_curr}",
                           report_curr, key="dl_vintage")
        st.markdown("---")
        gc1, gc2 = st.columns(2)
        v_sorted = df_v[df_v['Vintage'] != 'Total'].copy()
        vintages = v_sorted['Vintage'].astype(str).tolist()
        with gc1:
            fig_nav_v = go.Figure(go.Bar(
                x=vintages, y=v_sorted['% NAV'], marker_color='#4472C4',
                text=[f"<b>{v:.1f}%</b>" for v in v_sorted['% NAV']],
                textposition='outside', textfont=dict(size=14, color='#002060'),
            ))
            fig_nav_v.update_layout(
                title='% NAV por Vintage', height=450, plot_bgcolor='white', showlegend=False,
                yaxis=dict(showgrid=True, gridcolor='lightgrey', ticksuffix='%',
                           range=[0, v_sorted['% NAV'].max() * 1.25]),
                xaxis=dict(type='category'), margin=dict(t=60, b=40),
            )
            st.plotly_chart(fig_nav_v, use_container_width=True)
        with gc2:
            fig_comm_v = go.Figure(go.Bar(
                x=vintages, y=v_sorted['% Comm'], marker_color='#002060',
                text=[f"<b>{v:.1f}%</b>" for v in v_sorted['% Comm']],
                textposition='outside', textfont=dict(size=14, color='#002060'),
            ))
            fig_comm_v.update_layout(
                title='% Commitment por Vintage', height=450, plot_bgcolor='white', showlegend=False,
                yaxis=dict(showgrid=True, gridcolor='lightgrey', ticksuffix='%',
                           range=[0, v_sorted['% Comm'].max() * 1.25]),
                xaxis=dict(type='category'), margin=dict(t=60, b=40),
            )

    if tab_active("🎯 Estrategia"):
        fund_count_s = df_final.groupby('Strategy')['Fund'].count().reset_index().rename(columns={'Fund': '# Inv'})
        s_stats = s_stats.merge(fund_count_s, on='Strategy', how='left')

        STRAT_ORDER = ['Buyout','Growth Equity','Secondaries','Venture Capital','Fund of Funds','Single Co-Inv','Real Estate','Credit']
        s_stats['_order'] = s_stats['Strategy'].apply(lambda x: STRAT_ORDER.index(x) if x in STRAT_ORDER else 999)
        s_stats = s_stats.sort_values('_order').drop(columns='_order').reset_index(drop=True)

        # Totales globales para % columns
        total_comm  = s_stats['Commitment'].sum()
        total_nav   = s_stats['NAV'].sum()
        total_paid  = s_stats['Paid-In'].sum()
        total_funds = s_stats['# Inv'].sum()

        PE_STRATS = ['Buyout','Growth Equity','Secondaries','Venture Capital','Fund of Funds','Single Co-Inv']

        def build_row(label, mask, indent=False):
            sub = s_stats[mask]
            if sub.empty: return None
            comm  = sub['Commitment'].sum()
            paid  = sub['Paid-In'].sum()
            unf   = sub['Unfunded'].sum()
            dist  = sub['Distributed'].sum()
            nav   = sub['NAV'].sum()
            tv    = sub['Total Value'].sum()
            n_inv = int(sub['# Inv'].sum())
            irr   = calc_pooled_irr(df_final[df_final['Strategy'].isin(sub['Strategy'].tolist())])
            tvpi  = tv / paid if paid > 0 else 0
            dpi   = dist / paid if paid > 0 else 0
            dur   = calc_wav_duration(df_final[df_final['Strategy'].isin(sub['Strategy'].tolist())])
            return {
                'Strategy':         ('  ' if indent else '') + label,
                '# Inv':            n_inv,
                'Commit (MM)':      comm / 1e6,
                'Paid In (MM)':     paid / 1e6,
                'Unfunded (MM)':    unf  / 1e6,
                'Distributed (MM)': dist / 1e6,
                'NAV (MM)':         nav  / 1e6,
                'IRR':              irr,
                'TVPI':             tvpi,
                'DPI':              dpi,
                '% Comm':           comm / total_comm * 100 if total_comm > 0 else 0,
                '% NAV':            nav  / total_nav  * 100 if total_nav  > 0 else 0,
                '% Paid In':        paid / total_paid * 100 if total_paid > 0 else 0,
                'Duration (yrs)':   dur,
            }

        rows = []

        # Private Equity header
        pe_mask = s_stats['Strategy'].isin(PE_STRATS)
        pe_row  = build_row('Private Equity', pe_mask, indent=False)
        if pe_row: rows.append(pe_row)

        # PE sub-strategies
        for strat in PE_STRATS:
            mask = s_stats['Strategy'] == strat
            if not s_stats[mask].empty:
                r = build_row(strat, mask, indent=True)
                if r: rows.append(r)

        # Other top-level strategies
        for strat in ['Real Estate', 'Credit']:
            mask = s_stats['Strategy'] == strat
            if not s_stats[mask].empty:
                r = build_row(strat, mask, indent=False)
                if r: rows.append(r)

        # Remaining strategies not in the predefined lists
        defined = PE_STRATS + ['Real Estate', 'Credit']
        for strat in s_stats['Strategy'].tolist():
            if strat not in defined:
                mask = s_stats['Strategy'] == strat
                r = build_row(strat, mask, indent=False)
                if r: rows.append(r)

        # Total row
        rows.append({
            'Strategy':         'Total',
            '# Inv':            int(total_funds),
            'Commit (MM)':      total_comm / 1e6,
            'Paid In (MM)':     total_paid / 1e6,
            'Unfunded (MM)':    s_stats['Unfunded'].sum() / 1e6,
            'Distributed (MM)': s_stats['Distributed'].sum() / 1e6,
            'NAV (MM)':         total_nav / 1e6,
            'IRR':              g_irr,
            'TVPI':             s_stats['Total Value'].sum() / total_paid if total_paid > 0 else 0,
            'DPI':              s_stats['Distributed'].sum() / total_paid if total_paid > 0 else 0,
            '% Comm':           100.0,
            '% NAV':            100.0,
            '% Paid In':        100.0,
            'Duration (yrs)':   _dur_total,
        })

        df_hier = pd.DataFrame(rows)

        fmt_hier = {
            '# Inv':            '{:.0f}',
            'Commit (MM)':      '{:,.1f}',
            'Paid In (MM)':     '{:,.1f}',
            'Unfunded (MM)':    '{:,.1f}',
            'Distributed (MM)': '{:,.1f}',
            'NAV (MM)':         '{:,.1f}',
            'IRR':              '{:.1f}%',
            'TVPI':             '{:.2f}x',
            'DPI':              '{:.2f}x',
            '% Comm':           '{:.1f}%',
            '% NAV':            '{:.1f}%',
            '% Paid In':        '{:.1f}%',
            'Duration (yrs)':   '{:.1f}',
        }

        group_headers = {'Private Equity', 'Real Estate', 'Credit'}
        header_rows = {'Private Equity', 'Real Estate', 'Credit', 'Total'}
        def style_hier(row):
            name = row['Strategy'].strip()
            if name == 'Total':
                return ['font-weight:bold; background-color:#eef1f7'] * len(row)
            if name in group_headers:
                return ['font-weight:bold; background-color:#dce8f5'] * len(row)
            return ['color:#444444; font-size:12px'] * len(row)

        st.dataframe(
            df_hier.style.format(fmt_hier).apply(style_hier, axis=1),
            use_container_width=True,
            height=min(50 + len(df_hier) * 35, 600),
            hide_index=True,
        )
        excel_download_btn(df_hier, "Estrategia", f"estrategia_{report_curr}.xlsx",
                           "Estrategia", f"Portfolio por Estrategia — {report_curr}",
                           report_curr, key="dl_strat")
        st.markdown("---")
        s_sorted   = s_stats.reset_index(drop=True)
        strategies = s_sorted['Strategy'].tolist()
        irr_vals   = s_sorted['IRR %'].tolist()
        tvpi_vals  = s_sorted['TVPI'].tolist()
        fig_strat  = go.Figure()
        fig_strat.add_trace(go.Bar(
            name='IRR %', x=strategies, y=irr_vals, marker_color='#002060',
            text=[f"<b>{v:.1f}%</b>" for v in irr_vals], textposition='inside',
            textfont=dict(color='white', size=13), yaxis='y1',
        ))
        fig_strat.add_trace(go.Scatter(
            name='TVPI', x=strategies, y=tvpi_vals, mode='markers+text',
            marker=dict(symbol='diamond', size=14, color='#ED7D31'),
            text=[f"<b>{v:.2f}x</b>" for v in tvpi_vals], textposition='top center',
            textfont=dict(size=12, color='#ED7D31'), yaxis='y2',
        ))
        max_irr  = max(irr_vals)  if irr_vals  else 1
        max_tvpi = max(tvpi_vals) if tvpi_vals else 1
        fig_strat.update_layout(
            title=f'Performance by Strategy ({as_of_date_dt.strftime("%d/%m/%Y")})',
            height=480, plot_bgcolor='white', showlegend=True,
            legend=dict(orientation='h', y=1.08),
            yaxis=dict(showgrid=True, gridcolor='lightgrey', ticksuffix='%',
                       range=[0, max_irr * 1.5], title='IRR %'),
            yaxis2=dict(overlaying='y', side='right', showgrid=False,
                        range=[0, max_tvpi * 1.5], ticksuffix='x', title='TVPI'),
            xaxis=dict(type='category'), margin=dict(t=80, b=60),
        )
        st.plotly_chart(fig_strat, use_container_width=True)

    if tab_active("💼 Portfolio"):
        fmt_p = {c: '{:,.0f}' for c in money_cols if c in df_final.columns}
        fmt_p.update({'TVPI': '{:.2f}x', 'DPI': '{:.2f}x', 'IRR %': '{:.2f}%'})
        GRUPOS = {
            "🏦 Private Equity": ["Buyout","Secondaries","Growth Equity","Venture Capital","Fund of Funds"],
            "🎯 Co-Investments": ["Single Co-Inv"],
            "💳 Private Credit": ["Credit"],
            "🏢 Real Estate":    ["Real Estate"],
        }
        for grupo_nombre, estrategias in GRUPOS.items():
            df_grupo = df_final[df_final['Strategy'].isin(estrategias)].copy()
            if df_grupo.empty: continue
            df_grupo.index = range(1, len(df_grupo) + 1)
            st.markdown(f"### {grupo_nombre}")
            st.dataframe(df_grupo.style.format(fmt_p), use_container_width=True,
                         height=min(50 + len(df_grupo) * 36, 600))
            excel_download_btn(df_grupo, grupo_nombre.split()[-1],
                               f"portfolio_{grupo_nombre.split()[-1].lower()}_{report_curr}.xlsx",
                               grupo_nombre.split()[-1], f"{grupo_nombre} — {report_curr}",
                               report_curr, key=f"dl_port_{grupo_nombre}")
            t_paid_g = df_grupo['Paid-In'].sum(); t_dist_g = df_grupo['Distributed'].sum()
            t_nav_g  = df_grupo['NAV'].sum();     t_tv_g   = df_grupo['Total Value'].sum()
            t_comm_g = df_grupo['Commitment'].sum()
            try:
                flows_g = []
                for fn in df_grupo['Fund'].tolist():
                    fm = df_char[df_char['Fund'] == fn].iloc[0]; fc = fm['Currency']
                    ff = df_flows_raw[(df_flows_raw['Fund'] == fn) & (df_flows_raw['Date'] <= as_of_date_dt)].copy()
                    ff['Amt_Rep'] = [convert_amount(r['Amount'], fc, report_curr, fx_map.get(r['Date'].date(), fx_today)) for _, r in ff.iterrows()]
                    flows_g.append(ff[~ff['Type'].str.contains('NAV', case=False)][['Date','Amt_Rep']])
                agg_g = pd.concat(flows_g, ignore_index=True).groupby('Date')['Amt_Rep'].sum().reset_index()
                # Sumar NAV al último flujo si coincide con as_of, o agregar como fila nueva
                nav_date = pd.Timestamp(as_of_date_dt)
                if nav_date in agg_g['Date'].values:
                    agg_g.loc[agg_g['Date'] == nav_date, 'Amt_Rep'] += float(t_nav_g)
                    irr_dates   = agg_g['Date'].tolist()
                    irr_amounts = agg_g['Amt_Rep'].tolist()
                else:
                    irr_dates   = agg_g['Date'].tolist() + [nav_date]
                    irr_amounts = agg_g['Amt_Rep'].tolist() + [float(t_nav_g)]
                irr_g = xirr(irr_dates, irr_amounts) * 100
                if irr_g < -99 or irr_g > 500: irr_g = 0
            except: irr_g = 0
            tvpi_g = t_tv_g / t_paid_g if t_paid_g > 0 else 0
            c1,c2,c3,c4,c5 = st.columns(5)
            c1.metric("Commitment", f"{curr_sym}{t_comm_g/1e6:,.1f} M")
            c2.metric("Paid-In",    f"{curr_sym}{t_paid_g/1e6:,.1f} M")
            c3.metric("NAV",        f"{curr_sym}{t_nav_g/1e6:,.1f} M")
            c4.metric("Pooled IRR", f"{irr_g:.2f}%")
            c5.metric("TVPI",       f"{tvpi_g:.2f}x")
            st.markdown("---")

    CURR_OPTIONS = ["USD", "EUR", "GBP", "Local (moneda origen)"]
    CURR_KEYS    = {"USD":"USD","EUR":"EUR","GBP":"GBP","Local (moneda origen)":"Local"}
    DEFAULT_IDX  = {"USD":0,"EUR":1,"GBP":2}
    def curr_caption(cc):
        if cc == "Local":
            return "Flujos en moneda original de cada fondo — sin conversión FX."
        return f"Flujos convertidos a {cc} usando tipos de cambio históricos BCE."

    if tab_active("📉 IRR"):
        sel4 = st.radio("💱 Moneda de cálculo", CURR_OPTIONS, horizontal=True, key="curr_irr",
                        index=DEFAULT_IDX.get(report_curr, 0))
        cc4  = CURR_KEYS[sel4]
        st.caption(f"ℹ️ {curr_caption(cc4)}")
        irr_ev4, _, _, _ = calc_quarterly_evolutions(cc4)
        df_irr4 = pd.DataFrame(irr_ev4).set_index('Fund')
        df_irr4 = df_irr4[df_irr4.columns[::-1]]
        meta = df_final[['Fund','Strategy','Vintage']].set_index('Fund')
        df_irr4 = meta.join(df_irr4, how='right')
        q_cols4 = [c for c in df_irr4.columns if c not in ['Strategy','Vintage']]
        fmt_irr = {c: '{:.2f}%' for c in q_cols4}
        GRUPOS_EV = {
            "🏦 Private Equity": ["Buyout","Secondaries","Growth Equity","Venture Capital","Fund of Funds"],
            "🎯 Co-Investments": ["Single Co-Inv"],
            "💳 Private Credit": ["Credit"],
            "🏢 Real Estate":    ["Real Estate"],
        }
        for grupo_nombre, estrategias in GRUPOS_EV.items():
            df_g = df_irr4[df_irr4['Strategy'].isin(estrategias)].copy()
            if df_g.empty: continue
            df_g = df_g.sort_values(['Vintage','Fund'])
            df_g.index = range(1, len(df_g) + 1)
            st.markdown(f"### {grupo_nombre}")
            st.dataframe(df_g.style.format(fmt_irr, na_rep="-"),
                         use_container_width=True, height=min(50 + len(df_g)*36, 600))
            excel_download_btn(df_g.reset_index(), f"IRR {grupo_nombre.split()[-1]}",
                               f"irr_{grupo_nombre.split()[-1].lower()}_{cc4}.xlsx",
                               "IRR", f"IRR {grupo_nombre} — {cc4}",
                               report_curr, key=f"dl_irr_{grupo_nombre}")
            st.markdown("---")

    if tab_active("📈 TVPI"):
        sel5 = st.radio("💱 Moneda de cálculo", CURR_OPTIONS, horizontal=True, key="curr_tvpi",
                        index=DEFAULT_IDX.get(report_curr, 0))
        cc5  = CURR_KEYS[sel5]
        st.caption(f"ℹ️ {curr_caption(cc5)}")
        irr_ev5, tvpi_ev5, _, _ = calc_quarterly_evolutions(cc5)
        df_irr5  = pd.DataFrame(irr_ev5).set_index('Fund')
        df_irr5  = df_irr5[df_irr5.columns[::-1]]
        df_tvpi5 = pd.DataFrame(tvpi_ev5).set_index('Fund')
        df_tvpi5 = df_tvpi5[df_tvpi5.columns[::-1]]
        # Anular trimestres sin NAV (donde IRR es None)
        for col in df_tvpi5.columns:
            if col in df_irr5.columns:
                df_tvpi5[col] = df_tvpi5[col].where(df_irr5[col].notna(), other=None)
        meta = df_final[['Fund','Strategy','Vintage']].set_index('Fund')
        df_tvpi5 = meta.join(df_tvpi5, how='right')
        q_cols5 = [c for c in df_tvpi5.columns if c not in ['Strategy','Vintage']]
        fmt_tvpi = {c: '{:.2f}x' for c in q_cols5}
        for grupo_nombre, estrategias in GRUPOS_EV.items():
            df_g = df_tvpi5[df_tvpi5['Strategy'].isin(estrategias)].copy()
            if df_g.empty: continue
            df_g = df_g.sort_values(['Vintage','Fund'])
            df_g.index = range(1, len(df_g) + 1)
            st.markdown(f"### {grupo_nombre}")
            st.dataframe(df_g.style.format(fmt_tvpi, na_rep="-"),
                         use_container_width=True, height=min(50 + len(df_g)*36, 600))
            excel_download_btn(df_g.reset_index(), f"TVPI {grupo_nombre.split()[-1]}",
                               f"tvpi_{grupo_nombre.split()[-1].lower()}_{cc5}.xlsx",
                               "TVPI", f"TVPI {grupo_nombre} — {cc5}",
                               report_curr, key=f"dl_tvpi_{grupo_nombre}")
            st.markdown("---")

    if tab_active("💰 DPI"):
        sel6 = st.radio("💱 Moneda de cálculo", CURR_OPTIONS, horizontal=True, key="curr_dpi",
                        index=DEFAULT_IDX.get(report_curr, 0))
        cc6  = CURR_KEYS[sel6]
        st.caption(f"ℹ️ {curr_caption(cc6)}")
        _, _, dpi_ev6, _ = calc_quarterly_evolutions(cc6)
        df_dpi6 = pd.DataFrame(dpi_ev6).set_index('Fund')
        df_dpi6 = df_dpi6[df_dpi6.columns[::-1]]
        meta = df_final[['Fund','Strategy','Vintage']].set_index('Fund')
        df_dpi6 = meta.join(df_dpi6, how='right')
        q_cols6 = [c for c in df_dpi6.columns if c not in ['Strategy','Vintage']]
        fmt_dpi = {c: '{:.2f}x' for c in q_cols6}
        for grupo_nombre, estrategias in GRUPOS_EV.items():
            df_g = df_dpi6[df_dpi6['Strategy'].isin(estrategias)].copy()
            if df_g.empty: continue
            df_g = df_g.sort_values(['Vintage','Fund'])
            df_g.index = range(1, len(df_g) + 1)
            st.markdown(f"### {grupo_nombre}")
            st.dataframe(df_g.style.format(fmt_dpi, na_rep="-"),
                         use_container_width=True, height=min(50 + len(df_g)*36, 600))
            excel_download_btn(df_g.reset_index(), f"DPI {grupo_nombre.split()[-1]}",
                               f"dpi_{grupo_nombre.split()[-1].lower()}_{cc6}.xlsx",
                               "DPI", f"DPI {grupo_nombre} — {cc6}",
                               report_curr, key=f"dl_dpi_{grupo_nombre}")
            st.markdown("---")

    if tab_active("🔄 Rent."):
        sel7 = st.radio("💱 Moneda de cálculo", CURR_OPTIONS, horizontal=True, key="curr_rent",
                        index=DEFAULT_IDX.get(report_curr, 0))
        cc7  = CURR_KEYS[sel7]
        st.caption(f"ℹ️ {curr_caption(cc7)}")
        _, _, _, ret_ev7 = calc_quarterly_evolutions(cc7)
        df_ret7 = pd.DataFrame(ret_ev7).set_index('Fund')
        df_ret7 = df_ret7[df_ret7.columns[::-1]]
        meta = df_final[['Fund','Strategy','Vintage']].set_index('Fund')
        df_ret7 = meta.join(df_ret7, how='right')
        q_cols7 = [c for c in df_ret7.columns if c not in ['Strategy','Vintage']]
        fmt_ret = {c: '{:.2f}%' for c in q_cols7}
        for grupo_nombre, estrategias in GRUPOS_EV.items():
            df_g = df_ret7[df_ret7['Strategy'].isin(estrategias)].copy()
            if df_g.empty: continue
            df_g = df_g.sort_values(['Vintage','Fund'])
            df_g.index = range(1, len(df_g) + 1)
            st.markdown(f"### {grupo_nombre}")
            st.dataframe(df_g.style.format(fmt_ret, na_rep="N/A"),
                         use_container_width=True, height=min(50 + len(df_g)*36, 600))
            excel_download_btn(df_g.reset_index(), f"Rent. {grupo_nombre.split()[-1]}",
                               f"rent_{grupo_nombre.split()[-1].lower()}_{cc7}.xlsx",
                               "Rentabilidad", f"Rentabilidad {grupo_nombre} — {cc7}",
                               report_curr, key=f"dl_rent_{grupo_nombre}")
            st.markdown("---")

    if tab_active("💸 Cash Flows"):
        st.subheader(f"Cash Flows del Portfolio ({report_curr})")
        if not df_final.empty:
            # Selector trimestral / anual
            cf_view = st.radio("Vista", ["Trimestral", "Anual"], horizontal=True, key="cf_view")

            # Usar la fecha máxima de los datos (no la fecha de corte del filtro)
            cf_end_date = df_flows_raw['Date'].max()
            mo_end  = cf_end_date.month
            qm_end  = ((mo_end-1)//3+1)*3
            cf_end_qend = pd.Timestamp(cf_end_date.year, qm_end, 1) + pd.offsets.MonthEnd(0)

            # Siempre calcular trimestral como base
            q_dates_cf = pd.date_range(start=df_flows_raw['Date'].min(), end=cf_end_qend, freq='QE')
            cf_rows = []
            for q_d in q_dates_cf:
                qs = q_d - pd.tseries.offsets.QuarterEnd()
                total_calls = total_dists = total_nav = 0.0
                for f_name in df_final['Fund'].tolist():
                    f_mt = df_char[df_char['Fund'] == f_name].iloc[0]
                    f_fl = df_flows_raw[df_flows_raw['Fund'] == f_name].copy()
                    first_flow = f_fl['Date'].min() if not f_fl.empty else f_mt['Fecha Commitment']
                    if first_flow > q_d: continue
                    f_curr_ind = f_mt['Currency']
                    q_flows = f_fl[(f_fl['Date'] > qs) & (f_fl['Date'] <= q_d)]
                    for _, r in q_flows.iterrows():
                        amt_rep = convert_amount(r['Amount'], f_curr_ind, report_curr, fx_map.get(r['Date'].date(), fx_today))
                        if 'call' in r['Type'].lower(): total_calls += abs(amt_rep)
                        elif 'dist' in r['Type'].lower(): total_dists += amt_rep
                    nav_hist = f_fl[f_fl['Date'] <= q_d]
                    nav_entries_q = nav_hist[nav_hist['Type'].str.contains('NAV', case=False)].sort_values('Date')
                    if not nav_entries_q.empty:
                        ln_q   = nav_entries_q.iloc[-1]
                        later_q = nav_hist[nav_hist['Date'] > ln_q['Date']]
                        nav_loc_q = (float(ln_q['Amount'])
                                     + abs(float(later_q[later_q['Type'].str.contains('Call', case=False)]['Amount'].sum()))
                                     - float(later_q[later_q['Type'].str.contains('Dist', case=False)]['Amount'].sum()))
                        total_nav += convert_amount(nav_loc_q, f_curr_ind, report_curr, fx_map.get(q_d.date(), fx_today))
                    else:
                        calls_accum = abs(float(nav_hist[nav_hist['Type'].str.contains('Call', case=False)]['Amount'].sum()))
                        dists_accum = float(nav_hist[nav_hist['Type'].str.contains('Dist', case=False)]['Amount'].sum())
                        nav_cost = max(calls_accum - dists_accum, 0)
                        total_nav += convert_amount(nav_cost, f_curr_ind, report_curr, fx_map.get(q_d.date(), fx_today))
                cf_rows.append({
                    "Período": q_d.strftime('%d-%m-%Y'),
                    "Año": q_d.year,
                    "Capital Calls": -total_calls,
                    "Distribuciones": total_dists,
                    "NAV": total_nav,
                    "Net Cash Flow": total_dists - total_calls
                })

            df_cf_q = pd.DataFrame(cf_rows)

            if cf_view == "Anual":
                # Agregar por año: calls y dists suman, NAV toma el último del año
                df_cf_a = df_cf_q.groupby('Año').agg({
                    'Capital Calls': 'sum',
                    'Distribuciones': 'sum',
                    'Net Cash Flow': 'sum',
                    'NAV': 'last'   # NAV del último trimestre del año
                }).reset_index()
                df_cf_a['Año'] = df_cf_a['Año'].astype(str)
                df_cf = df_cf_a.set_index('Año').iloc[::-1]  # más reciente primero
                periodo_label = "Año"
                titulo_grafico = f"Capital Calls, Distribuciones y NAV por Año ({report_curr})"
                titulo_tabla   = "#### Detalle por año"
                dl_key = "dl_cf_a"
                dl_file = f"cashflows_anual_{report_curr}.xlsx"
            else:
                df_cf = df_cf_q.set_index('Período').drop(columns='Año').iloc[::-1]
                periodo_label = "Trimestre"
                titulo_grafico = f"Capital Calls, Distribuciones y NAV por Trimestre ({report_curr})"
                titulo_tabla   = "#### Detalle por trimestre"
                dl_key = "dl_cf_q"
                dl_file = f"cashflows_trimestral_{report_curr}.xlsx"

            # ── Gráfico ───────────────────────────────────────────────────────
            # Para el gráfico usar orden cronológico
            df_cf_chart = df_cf.iloc[::-1]
            fig_cf = go.Figure()
            fig_cf.add_trace(go.Bar(name="Capital Calls", x=df_cf_chart.index, y=df_cf_chart["Capital Calls"],
                                    marker_color="#ED7D31",
                                    text=[f"{v/1e6:,.1f}" for v in df_cf_chart["Capital Calls"]], textposition="outside"))
            fig_cf.add_trace(go.Bar(name="Distribuciones", x=df_cf_chart.index, y=df_cf_chart["Distribuciones"],
                                    marker_color="#92D050",
                                    text=[f"{v/1e6:,.1f}" for v in df_cf_chart["Distribuciones"]], textposition="outside"))
            fig_cf.add_trace(go.Scatter(name="NAV", x=df_cf_chart.index, y=df_cf_chart["NAV"],
                                        mode="lines+markers", line=dict(color="#FFC000", width=2),
                                        marker=dict(size=6), yaxis="y2"))
            fig_cf.update_layout(
                barmode="relative", height=500, plot_bgcolor="white",
                yaxis=dict(title=f"Cash Flow ({report_curr} M)", showgrid=True, gridcolor="lightgrey", tickformat=",.0f"),
                yaxis2=dict(title=f"NAV ({report_curr})", overlaying="y", side="right", showgrid=False, tickformat=",.0f"),
                legend=dict(orientation="h", y=1.08),
                title=titulo_grafico,
                xaxis=dict(tickangle=-45),
            )
            st.plotly_chart(fig_cf, use_container_width=True)

            # ── Debug (solo trimestral) ───────────────────────────────────────
            if cf_view == "Trimestral":
                with st.expander("🔍 Debug: desglose por fondo en un trimestre", expanded=False):
                    q_debug = st.selectbox("Trimestre a inspeccionar",
                                           options=list(df_cf.index), index=0, key="cf_debug_q")
                    q_d_dbg = pd.to_datetime(q_debug, dayfirst=True)
                    qs_dbg  = q_d_dbg - pd.tseries.offsets.QuarterEnd()
                    debug_rows = []
                    for f_name in df_final['Fund'].tolist():
                        f_mt = df_char[df_char['Fund'] == f_name].iloc[0]
                        f_curr_ind = f_mt['Currency']
                        f_fl_dbg = df_flows_raw[df_flows_raw['Fund'] == f_name].copy()
                        first_flow_dbg = f_fl_dbg['Date'].min() if not f_fl_dbg.empty else f_mt['Fecha Commitment']
                        if first_flow_dbg > q_d_dbg: continue
                        q_fl = f_fl_dbg[(f_fl_dbg['Date'] > qs_dbg) & (f_fl_dbg['Date'] <= q_d_dbg)]
                        for _, r in q_fl.iterrows():
                            amt_rep = convert_amount(r['Amount'], f_curr_ind, report_curr,
                                                     fx_map.get(r['Date'].date(), fx_today))
                            debug_rows.append({
                                'Fund': f_name, 'Currency': f_curr_ind,
                                'Date': r['Date'].strftime('%d/%m/%Y'),
                                'Type': r['Type'],
                                'Amount (orig)': r['Amount'],
                                f'Amount ({report_curr})': amt_rep,
                            })
                    if debug_rows:
                        df_dbg = pd.DataFrame(debug_rows)
                        st.dataframe(df_dbg.style.format({
                            'Amount (orig)': '{:,.0f}',
                            f'Amount ({report_curr})': '{:,.0f}'
                        }), use_container_width=True)
                        calls_dbg = df_dbg[df_dbg['Type'].str.contains('Call', case=False)][f'Amount ({report_curr})'].sum()
                        dists_dbg = df_dbg[df_dbg['Type'].str.contains('Dist', case=False)][f'Amount ({report_curr})'].sum()
                        st.write(f"**Total Calls:** {calls_dbg:,.0f} | **Total Dists:** {dists_dbg:,.0f}")
                    else:
                        st.info("Sin flujos en este trimestre.")

            # ── Tabla ─────────────────────────────────────────────────────────
            st.markdown(titulo_tabla)
            fmt_cf = {"Capital Calls":"{:,.0f}","Distribuciones":"{:,.0f}","NAV":"{:,.0f}","Net Cash Flow":"{:,.0f}"}
            def color_netcf(val):
                return f"background-color: {'#d4edda' if val >= 0 else '#f8d7da'}"
            st.dataframe(df_cf.style.format(fmt_cf).map(color_netcf, subset=["Net Cash Flow"]),
                         use_container_width=True)
            excel_download_btn(df_cf.reset_index(), "Cash Flows", dl_file, "Cash Flows",
                               f"Cash Flows — {report_curr}", report_curr, key=dl_key)
            st.markdown("---")
            c1,c2,c3,c4 = st.columns(4)
            c1.metric("Total Capital Calls",  f"{curr_sym}{abs(df_cf['Capital Calls'].sum())/1e6:,.1f} M")
            c2.metric("Total Distribuciones", f"{curr_sym}{df_cf['Distribuciones'].sum()/1e6:,.1f} M")
            c3.metric("NAV Actual",           f"{curr_sym}{t_nav/1e6:,.1f} M")
            c4.metric("Net Cash Flow Total",  f"{curr_sym}{df_cf['Net Cash Flow'].sum()/1e6:,.1f} M")

    if tab_active("📆 Commitment Pace"):
        st.subheader(f"Commitment Pace ({report_curr})")
        CO_INV_STRATS = ["Single Co-Inv"]
        df_pace = df_char_filt.copy()
        df_pace['Commitment_Rep'] = df_pace.apply(
            lambda r: convert_amount(r['Commitment'], r['Currency'], report_curr,
                                     fx_map.get(r['Fecha Commitment'].date(), fx_today)), axis=1)
        df_pace['Año Commitment'] = df_pace['Fecha Commitment'].dt.year
        df_pace['Es CoInv']       = df_pace['Strategy'].isin(CO_INV_STRATS)
        def build_pace_chart(df_src, group_col, title):
            grp = df_src.groupby([group_col,'Es CoInv'])['Commitment_Rep'].sum().reset_index()
            years = sorted(df_src[group_col].unique())
            fondos_y=[];coinv_y=[];total_y=[];pct_coinv=[]
            for y in years:
                sub   = grp[grp[group_col]==y]
                f_val = float(sub[~sub['Es CoInv']]['Commitment_Rep'].sum())
                c_val = float(sub[sub['Es CoInv']]['Commitment_Rep'].sum())
                tot   = f_val + c_val
                fondos_y.append(f_val);coinv_y.append(c_val);total_y.append(tot)
                pct_coinv.append(c_val/tot*100 if tot>0 else 0)
            years_str = [str(y) for y in years]
            fig = go.Figure()
            fig.add_trace(go.Bar(name='Fondos', x=years_str, y=[v/1e6 for v in fondos_y],
                                 marker_color='#002060', text=[f"<b>{v/1e6:.1f}</b>" for v in fondos_y],
                                 textposition='inside', textfont=dict(color='white',size=12), yaxis='y1'))
            fig.add_trace(go.Bar(name='Single Co-Inv', x=years_str, y=[v/1e6 for v in coinv_y],
                                 marker_color='#ED7D31',
                                 text=[f"<b>{v/1e6:.1f}</b>" if v>0 else '' for v in coinv_y],
                                 textposition='inside', textfont=dict(color='white',size=12), yaxis='y1'))
            fig.add_trace(go.Scatter(x=years_str, y=[v/1e6 for v in total_y], mode='text',
                                     text=[f"<b>{v/1e6:.1f}</b>" for v in total_y],
                                     textposition='top center', textfont=dict(size=13,color='#002060'),
                                     showlegend=False, yaxis='y1'))
            fig.add_trace(go.Scatter(name='% Co-Inv', x=years_str, y=pct_coinv,
                                     mode='lines+markers+text', line=dict(color='#4472C4',width=2),
                                     marker=dict(size=7),
                                     text=[f"<b>{v:.0f}%</b>" if v>0 else '-' for v in pct_coinv],
                                     textposition='top center', textfont=dict(size=12,color='#4472C4'),
                                     yaxis='y2'))
            max_mm = max(total_y)/1e6 if total_y else 1
            fig.update_layout(barmode='stack', title=title, height=480, plot_bgcolor='white',
                              legend=dict(orientation='h',y=1.08),
                              yaxis=dict(title=f'Commitment ({report_curr} MM)',showgrid=True,gridcolor='lightgrey',range=[0,max_mm*1.35]),
                              yaxis2=dict(title='% Co-Inv',overlaying='y',side='right',showgrid=False,range=[-110,110],ticksuffix='%'),
                              xaxis=dict(type='category'), margin=dict(t=80,b=40))
            return fig

        def build_pace_table(df_src, group_col):
            """Tabla % compromiso por estrategia: años en filas (más nuevo primero), estrategias en columnas."""
            grp = df_src.groupby([group_col, 'Strategy'])['Commitment_Rep'].sum().reset_index()
            years = sorted(df_src[group_col].unique(), reverse=True)  # más nuevo primero
            strategies = sorted(df_src['Strategy'].unique())
            rows = []
            for y in years:
                total_y = grp[grp[group_col] == y]['Commitment_Rep'].sum()
                row = {'Año': str(int(y))}
                for strat in strategies:
                    strat_y = float(grp[(grp[group_col] == y) & (grp['Strategy'] == strat)]['Commitment_Rep'].sum())
                    row[strat] = (strat_y / total_y * 100) if total_y > 0 else 0.0
                row['Total'] = 100.0 if total_y > 0 else 0.0
                rows.append(row)
            df_tbl = pd.DataFrame(rows).set_index('Año')
            fmt = {c: '{:.1f}%' for c in df_tbl.columns}
            def highlight_total_col(row):
                return ['font-weight:bold; background-color:#eef1f7' if c == 'Total'
                        else '' for c in row.index]
            st.dataframe(
                df_tbl.style.format(fmt).apply(highlight_total_col, axis=1),
                use_container_width=True,
                height=min(50 + len(df_tbl) * 35, 400)
            )

        st.markdown("#### Por Vintage Year")
        st.plotly_chart(build_pace_chart(df_pace,'Vintage',f'Commitment por Vintage Year ({report_curr} MM)'),
                        use_container_width=True)
        st.markdown("##### % Compromiso por Estrategia — Vintage Year")
        build_pace_table(df_pace, 'Vintage')
        st.markdown("---")
        st.markdown("#### Por Año de Fecha Commitment")
        st.plotly_chart(build_pace_chart(df_pace,'Año Commitment',f'Commitment por Año de Inversión ({report_curr} MM)'),
                        use_container_width=True)
        st.markdown("##### % Compromiso por Estrategia — Año Commitment")
        build_pace_table(df_pace, 'Año Commitment')

    # TAB 10 — POINT IN TIME
    # =========================================================================
    if tab_active("📍 Point in Time"):
        st.markdown("### 📍 Point in Time — Rendimiento por Período")
        st.caption("Calcula Money-Weighted Return (MWR/IRR) y Time-Weighted Return (TWR) "
                   "entre dos fechas seleccionadas, por estrategia y total.")

        # ── Selectores de fecha ───────────────────────────────────────────────
        pit_col1, pit_col2 = st.columns(2)
        with pit_col1:
            min_date_pit = df_flows_raw['Date'].min().date()
            pit_start = st.date_input("📅 Fecha inicio", key="pit_start",
                                       value=min_date_pit,
                                       min_value=min_date_pit)
        with pit_col2:
            pit_end = st.date_input("📅 Fecha fin", key="pit_end",
                                     value=as_of_date,
                                     max_value=as_of_date)

        pit_start_dt = pd.Timestamp(pit_start)
        pit_end_dt   = pd.Timestamp(pit_end)

        if pit_start_dt >= pit_end_dt:
            st.error("La fecha de inicio debe ser anterior a la fecha fin.")
        elif df_final.empty:
            st.info("Sin datos en el portfolio filtrado.")
        else:
            # ── Funciones de cálculo ──────────────────────────────────────────

            def get_nav_at(fund_name, f_curr, date_ts):
                """NAV de un fondo en una fecha específica (interpolado desde cashflows)."""
                f_fl = df_flows_raw[
                    (df_flows_raw['Fund'] == fund_name) &
                    (df_flows_raw['Date'] <= date_ts)
                ].copy()
                nav_entries = f_fl[f_fl['Type'].str.contains('NAV', case=False)].sort_values('Date')
                if not nav_entries.empty:
                    ln = nav_entries.iloc[-1]
                    later = f_fl[f_fl['Date'] > ln['Date']]
                    nav_loc = (float(ln['Amount'])
                               + abs(float(later[later['Type'].str.contains('Call', case=False)]['Amount'].sum()))
                               - float(later[later['Type'].str.contains('Dist', case=False)]['Amount'].sum()))
                    return convert_amount(nav_loc, f_curr, report_curr,
                                         fx_map.get(date_ts.date(), fx_today))
                # Sin NAV: usar paid-in acumulado como proxy
                calls = abs(f_fl[f_fl['Type'].str.contains('Call', case=False)]['Amount'].sum())
                dists = f_fl[f_fl['Type'].str.contains('Dist', case=False)]['Amount'].sum()
                return convert_amount(max(calls - dists, 0), f_curr, report_curr,
                                      fx_map.get(date_ts.date(), fx_today))

            def calc_mwr_period(fund_list, start_dt, end_dt):
                """
                Money-Weighted Return (IRR) entre start y end.
                Devuelve (mwr_ann_total, mwr_anualizado).
                El xirr ya es una tasa anualizada por definición.
                El total se obtiene: (1 + mwr_anual)^(días/365) - 1
                """
                all_flows = []
                nav_start_total = 0.0
                nav_end_total   = 0.0

                for fund in fund_list:
                    f_meta = df_char[df_char['Fund'] == fund]
                    if f_meta.empty: continue
                    f_meta = f_meta.iloc[0]
                    f_curr = f_meta['Currency']

                    nav_s = get_nav_at(fund, f_curr, start_dt)
                    nav_e = get_nav_at(fund, f_curr, end_dt)
                    nav_start_total += nav_s
                    nav_end_total   += nav_e

                    mid_flows = df_flows_raw[
                        (df_flows_raw['Fund'] == fund) &
                        (df_flows_raw['Date'] > start_dt) &
                        (df_flows_raw['Date'] <= end_dt) &
                        (~df_flows_raw['Type'].str.contains('NAV', case=False))
                    ].copy()
                    for _, r in mid_flows.iterrows():
                        amt = convert_amount(r['Amount'], f_curr, report_curr,
                                             fx_map.get(r['Date'].date(), fx_today))
                        all_flows.append({'Date': r['Date'], 'Amt': amt})

                if nav_start_total <= 0 and not all_flows:
                    return 0.0, 0.0

                flow_df  = pd.DataFrame(all_flows) if all_flows else pd.DataFrame(columns=['Date','Amt'])
                flow_agg = flow_df.groupby('Date')['Amt'].sum().reset_index() if not flow_df.empty \
                           else pd.DataFrame(columns=['Date','Amt'])

                dates   = [start_dt] + flow_agg['Date'].tolist() + [end_dt]
                amounts = [-nav_start_total] + flow_agg['Amt'].tolist() + [nav_end_total]

                from collections import defaultdict
                merged = defaultdict(float)
                for d, a in zip(dates, amounts):
                    merged[d] += a
                dates_m   = sorted(merged.keys())
                amounts_m = [merged[d] for d in dates_m]

                if len(dates_m) < 2:
                    return 0.0, 0.0
                try:
                    mwr_ann = xirr(dates_m, amounts_m) * 100   # ya es anualizado
                    if mwr_ann <= -99:
                        return 0.0, 0.0
                    days = (end_dt - start_dt).days
                    if days <= 0:
                        return mwr_ann, mwr_ann
                    # Total del período: (1 + mwr_ann/100)^(días/365) - 1
                    mwr_ann_total_pct = ((1 + mwr_ann / 100) ** (days / 365.25) - 1) * 100
                    return mwr_ann_total_pct, mwr_ann
                except:
                    return 0.0, 0.0

            def calc_twr_period(fund_list, start_dt, end_dt):
                """
                Time-Weighted Return entre start y end.
                Devuelve (twr_ann_total_pct, twr_anualizado_pct).
                """
                q_dates    = pd.date_range(start=start_dt, end=end_dt, freq='QE')
                eval_dates = sorted(set([start_dt] + q_dates.tolist() + [end_dt]))
                clean_dates = [eval_dates[0]]
                for d in eval_dates[1:]:
                    if (d - clean_dates[-1]).days >= 1:
                        clean_dates.append(d)

                if len(clean_dates) < 2:
                    return 0.0, 0.0

                product = 1.0
                for i in range(len(clean_dates) - 1):
                    t0 = clean_dates[i]
                    t1 = clean_dates[i + 1]
                    nav_t0 = nav_t1 = calls_period = dists_period = 0.0

                    for fund in fund_list:
                        f_meta = df_char[df_char['Fund'] == fund]
                        if f_meta.empty: continue
                        f_meta = f_meta.iloc[0]
                        f_curr = f_meta['Currency']
                        nav_t0 += get_nav_at(fund, f_curr, t0)
                        nav_t1 += get_nav_at(fund, f_curr, t1)
                        sub = df_flows_raw[
                            (df_flows_raw['Fund'] == fund) &
                            (df_flows_raw['Date'] > t0) &
                            (df_flows_raw['Date'] <= t1) &
                            (~df_flows_raw['Type'].str.contains('NAV', case=False))
                        ]
                        for _, r in sub.iterrows():
                            amt = convert_amount(r['Amount'], f_curr, report_curr,
                                                 fx_map.get(r['Date'].date(), fx_today))
                            if r['Type'].lower().find('call') >= 0:
                                calls_period += abs(amt)
                            elif r['Type'].lower().find('dist') >= 0:
                                dists_period += amt

                    denom = nav_t0 + calls_period
                    if denom <= 0:
                        continue
                    product *= (nav_t1 + dists_period) / denom

                twr_acum = product - 1
                days = (end_dt - start_dt).days
                if days <= 0:
                    return twr_acum * 100, twr_acum * 100
                twr_ann_total_pct = twr_acum * 100
                twr_ann_pct   = ((1 + twr_acum) ** (365.25 / days) - 1) * 100
                if twr_ann_pct <= -9999:
                    return 0.0, 0.0
                return twr_ann_total_pct, twr_ann_pct

            # ── Calcular por estrategia ───────────────────────────────────────
            STRAT_ORDER_PIT = ['Buyout','Growth Equity','Secondaries','Venture Capital',
                               'Fund of Funds','Single Co-Inv','Real Estate','Credit']

            pit_rows = []
            strategies_in_portfolio = df_final['Strategy'].unique().tolist()
            strategies_ordered = [s for s in STRAT_ORDER_PIT if s in strategies_in_portfolio] + \
                                  [s for s in strategies_in_portfolio if s not in STRAT_ORDER_PIT]

            with st.spinner("Calculando rendimientos del período..."):
                for strat in strategies_ordered:
                    funds_strat = df_final[df_final['Strategy'] == strat]['Fund'].tolist()
                    if not funds_strat: continue

                    nav_s = sum(get_nav_at(f, df_char[df_char['Fund']==f].iloc[0]['Currency'],
                                           pit_start_dt) for f in funds_strat
                               if not df_char[df_char['Fund']==f].empty)
                    nav_e = sum(get_nav_at(f, df_char[df_char['Fund']==f].iloc[0]['Currency'],
                                           pit_end_dt) for f in funds_strat
                               if not df_char[df_char['Fund']==f].empty)

                    # Flujos del período
                    calls_p = dists_p = 0.0
                    for fund in funds_strat:
                        fm = df_char[df_char['Fund']==fund]
                        if fm.empty: continue
                        f_curr = fm.iloc[0]['Currency']
                        mid = df_flows_raw[
                            (df_flows_raw['Fund']==fund) &
                            (df_flows_raw['Date'] > pit_start_dt) &
                            (df_flows_raw['Date'] <= pit_end_dt) &
                            (~df_flows_raw['Type'].str.contains('NAV', case=False))
                        ]
                        for _, r in mid.iterrows():
                            amt = convert_amount(r['Amount'], f_curr, report_curr,
                                                 fx_map.get(r['Date'].date(), fx_today))
                            if 'call' in r['Type'].lower(): calls_p += abs(amt)
                            elif 'dist' in r['Type'].lower(): dists_p += amt

                    mwr_tot, mwr_ann = calc_mwr_period(funds_strat, pit_start_dt, pit_end_dt)
                    twr_tot, twr_ann = calc_twr_period(funds_strat, pit_start_dt, pit_end_dt)
                    n_funds_s = len(funds_strat)

                    pit_rows.append({
                        'Estrategia':       strat,
                        'N° Fondos':        n_funds_s,
                        'NAV Inicio':       nav_s,
                        'Capital Calls':    calls_p,
                        'Distribuciones':   dists_p,
                        'NAV Fin':          nav_e,
                        'MWR Total':        mwr_tot,
                        'MWR Anualizado':   mwr_ann,
                        'TWR Total':        twr_tot,
                        'TWR Anualizado':   twr_ann,
                    })

                # Total portfolio
                all_funds = df_final['Fund'].tolist()
                mwr_tot_total, mwr_ann_total = calc_mwr_period(all_funds, pit_start_dt, pit_end_dt)
                twr_tot_total, twr_ann_total = calc_twr_period(all_funds, pit_start_dt, pit_end_dt)
                nav_s_total = sum(
                    get_nav_at(f, df_char[df_char['Fund']==f].iloc[0]['Currency'], pit_start_dt)
                    for f in all_funds if not df_char[df_char['Fund']==f].empty)
                nav_e_total = sum(
                    get_nav_at(f, df_char[df_char['Fund']==f].iloc[0]['Currency'], pit_end_dt)
                    for f in all_funds if not df_char[df_char['Fund']==f].empty)
                calls_total = sum(r['Capital Calls'] for r in pit_rows)
                dists_total = sum(r['Distribuciones'] for r in pit_rows)

            df_pit = pd.DataFrame(pit_rows)

            # ── Métricas globales ─────────────────────────────────────────────
            st.markdown(f"""
            <div style='background:#f0f7ff;border-left:3px solid #1a5fd4;
            padding:10px 16px;border-radius:4px;font-size:12px;margin:12px 0 16px'>
            📅 <b>Período:</b> {pit_start.strftime('%d/%m/%Y')} → {pit_end.strftime('%d/%m/%Y')}
            &nbsp;&nbsp;|&nbsp;&nbsp;
            <b>NAV Inicio:</b> {curr_sym}{nav_s_total/1e6:,.2f}M
            &nbsp;&nbsp;|&nbsp;&nbsp;
            <b>NAV Fin:</b> {curr_sym}{nav_e_total/1e6:,.2f}M
            &nbsp;&nbsp;|&nbsp;&nbsp;
            <b>Calls:</b> {curr_sym}{calls_total/1e6:,.2f}M
            &nbsp;&nbsp;|&nbsp;&nbsp;
            <b>Distribuciones:</b> {curr_sym}{dists_total/1e6:,.2f}M
            </div>
            """, unsafe_allow_html=True)

            km1, km2, km3 = st.columns(3)
            km1, km2 = st.columns(2)
            with km1:
                st.markdown("**MWR (Money-Weighted)**")
                ma1, ma2 = st.columns(2)
                ma1.metric("Total período", f"{mwr_tot_total:.2f}%",
                            help="Retorno total acumulado del período")
                ma2.metric("Anualizado (IRR)", f"{mwr_ann_total:.2f}%",
                            help="IRR anualizado: trata NAV inicio como inversión y NAV fin como retorno")
            with km2:
                st.markdown("**TWR (Time-Weighted)**")
                ta1, ta2 = st.columns(2)
                ta1.metric("Total período", f"{twr_tot_total:.2f}%",
                            help="Retorno total acumulado sin efecto de timing")
                ta2.metric("Anualizado", f"{twr_ann_total:.2f}%",
                            help="TWR anualizado: (1 + TWR_total)^(365/días) - 1")
            nav_change = nav_e_total - nav_s_total + dists_total - calls_total
            st.markdown("---")
            st.metric("Ganancia del Período", f"{curr_sym}{nav_change/1e6:,.2f}M",
                       help="(NAV Fin + Distribuciones) - (NAV Inicio + Capital Calls)")

            st.markdown("---")

            # ── Tabla agregada por estrategia ────────────────────────────────
            st.markdown("#### Rendimiento por Estrategia")

            fmt_pit = {
                'N° Fondos':      '{:.0f}',
                'NAV Inicio':     '{:,.0f}',
                'Capital Calls':  '{:,.0f}',
                'Distribuciones': '{:,.0f}',
                'NAV Fin':        '{:,.0f}',
                'MWR Total':      '{:.2f}%',
                'MWR Anualizado': '{:.2f}%',
                'TWR Total':      '{:.2f}%',
                'TWR Anualizado': '{:.2f}%',
            }

            def color_returns(val):
                if isinstance(val, (int, float)):
                    if val > 0:   return 'color:#00703c; font-weight:600'
                    elif val < 0: return 'color:#c00000; font-weight:600'
                return ''

            def highlight_pit_total(row):
                if row.name == 'Total':
                    return ['font-weight:bold; background-color:#eef1f7'] * len(row)
                return [''] * len(row)

            if not df_pit.empty:
                # Tabla completa con fila TOTAL
                total_pit_row = pd.DataFrame([{
                    'Estrategia':       'TOTAL',
                    'N° Fondos':        df_final['Fund'].nunique(),
                    'NAV Inicio':       nav_s_total,
                    'Capital Calls':    calls_total,
                    'Distribuciones':   dists_total,
                    'NAV Fin':          nav_e_total,
                    'MWR Total':        mwr_tot_total,
                    'MWR Anualizado':   mwr_ann_total,
                    'TWR Total':        twr_tot_total,
                    'TWR Anualizado':   twr_ann_total,
                }])
                df_pit_display = pd.concat([df_pit, total_pit_row], ignore_index=True)
                df_pit_display.index = list(range(1, len(df_pit)+1)) + ['Total']

                st.dataframe(
                    df_pit_display.style
                        .format(fmt_pit)
                        .apply(highlight_pit_total, axis=1)
                        .map(color_returns, subset=['MWR Total','MWR Anualizado','TWR Total','TWR Anualizado']),
                    use_container_width=True,
                    height=min(60 + len(df_pit_display) * 36, 500)
                )
                excel_download_btn(df_pit_display.reset_index(), "Point in Time",
                                   f"point_in_time_{report_curr}.xlsx", "Point in Time",
                                   f"Rendimiento {pit_start.strftime('%d/%m/%Y')} → {pit_end.strftime('%d/%m/%Y')}",
                                   report_curr, key="dl_pit")

                # ── Detalle por estrategia (expandible) ───────────────────────
                st.markdown("---")
                st.markdown("#### Detalle por Estrategia")
                st.caption("Haz clic en una estrategia para ver sus fondos individuales.")

                fmt_f = {
                    'NAV Inicio':     '{:,.0f}',
                    'Capital Calls':  '{:,.0f}',
                    'Distribuciones': '{:,.0f}',
                    'NAV Fin':        '{:,.0f}',
                    'MWR Total':      '{:.2f}%',
                    'MWR Anualizado': '{:.2f}%',
                    'TWR Total':      '{:.2f}%',
                    'TWR Anualizado': '{:.2f}%',
                }

                for _, srow in df_pit.iterrows():
                    strat      = srow['Estrategia']
                    funds_list = df_final[df_final['Strategy'] == strat]['Fund'].tolist()
                    mwr_color  = '🟢' if srow['MWR Total'] > 0 else '🔴'

                    with st.expander(
                        f"{mwr_color} **{strat}** — "
                        f"MWR: {srow['MWR Total']:.2f}% ({srow['MWR Anualizado']:.2f}% ann)  |  "
                        f"TWR: {srow['TWR Total']:.2f}% ({srow['TWR Anualizado']:.2f}% ann)  |  "
                        f"{int(srow['N° Fondos'])} fondos",
                        expanded=False
                    ):
                        fund_rows = []
                        for fund in funds_list:
                            fm = df_char[df_char['Fund'] == fund]
                            if fm.empty: continue
                            f_curr = fm.iloc[0]['Currency']
                            gp     = fm.iloc[0]['GP']
                            nav_s_f = get_nav_at(fund, f_curr, pit_start_dt)
                            nav_e_f = get_nav_at(fund, f_curr, pit_end_dt)
                            calls_f = dists_f = 0.0
                            mid_f = df_flows_raw[
                                (df_flows_raw['Fund'] == fund) &
                                (df_flows_raw['Date'] > pit_start_dt) &
                                (df_flows_raw['Date'] <= pit_end_dt) &
                                (~df_flows_raw['Type'].str.contains('NAV', case=False))
                            ]
                            for _, r in mid_f.iterrows():
                                amt = convert_amount(r['Amount'], f_curr, report_curr,
                                                     fx_map.get(r['Date'].date(), fx_today))
                                if 'call' in r['Type'].lower(): calls_f += abs(amt)
                                elif 'dist' in r['Type'].lower(): dists_f += amt

                            mwr_f_tot, mwr_f_ann = calc_mwr_period([fund], pit_start_dt, pit_end_dt)
                            twr_f_tot, twr_f_ann = calc_twr_period([fund], pit_start_dt, pit_end_dt)
                            fund_rows.append({
                                'Fund':           fund,
                                'GP':             gp,
                                'NAV Inicio':     nav_s_f,
                                'Capital Calls':  calls_f,
                                'Distribuciones': dists_f,
                                'NAV Fin':        nav_e_f,
                                'MWR Total':      mwr_f_tot,
                                'MWR Anualizado': mwr_f_ann,
                                'TWR Total':      twr_f_tot,
                                'TWR Anualizado': twr_f_ann,
                            })

                        df_funds_pit = pd.DataFrame(fund_rows)
                        if not df_funds_pit.empty:
                            df_funds_pit.index = range(1, len(df_funds_pit) + 1)
                            st.dataframe(
                                df_funds_pit.style
                                    .format(fmt_f)
                                    .map(color_returns,
                                              subset=['MWR Total','MWR Anualizado','TWR Total','TWR Anualizado']),
                                use_container_width=True,
                                height=min(50 + len(df_funds_pit) * 35, 420)
                            )

            # ── Gráfico comparativo MWR vs TWR ────────────────────────────────
            st.markdown("---")
            st.markdown("#### MWR vs TWR por Estrategia")
            if not df_pit.empty:
                strats_pit = df_pit['Estrategia'].tolist()
                mwr_vals   = df_pit['MWR Anualizado'].tolist()
                twr_vals   = df_pit['TWR Anualizado'].tolist()

                fig_pit = go.Figure()
                fig_pit.add_trace(go.Bar(
                    name='MWR Anualizado', x=strats_pit, y=mwr_vals,
                    marker_color='#002060',
                    text=[f"<b>{v:.1f}%</b>" for v in mwr_vals],
                    textposition='outside', textfont=dict(size=11, color='#002060'),
                ))
                fig_pit.add_trace(go.Bar(
                    name='TWR Anualizado', x=strats_pit, y=twr_vals,
                    marker_color='#4472C4',
                    text=[f"<b>{v:.1f}%</b>" for v in twr_vals],
                    textposition='outside', textfont=dict(size=11, color='#4472C4'),
                ))
                # Línea de portfolio total
                fig_pit.add_hline(y=mwr_ann_total, line_dash='dash', line_color='#ED7D31',
                                   annotation_text=f"MWR Total: {mwr_ann_total:.1f}%",
                                   annotation_position="top right")
                fig_pit.add_hline(y=twr_ann_total, line_dash='dot', line_color='#92D050',
                                   annotation_text=f"TWR Total: {twr_ann_total:.1f}%",
                                   annotation_position="bottom right")
                max_abs = max(abs(v) for v in mwr_vals + twr_vals + [mwr_ann_total, twr_ann_total]) * 1.4
                fig_pit.update_layout(
                    barmode='group', height=460, plot_bgcolor='white',
                    title=f'MWR vs TWR por Estrategia — {pit_start.strftime("%d/%m/%Y")} a {pit_end.strftime("%d/%m/%Y")}',
                    yaxis=dict(showgrid=True, gridcolor='#eef1f7', ticksuffix='%',
                               zeroline=True, zerolinecolor='#aaa', zerolinewidth=1.5,
                               range=[-max_abs, max_abs]),
                    legend=dict(orientation='h', y=1.08),
                    xaxis=dict(type='category'),
                    margin=dict(t=80, b=60),
                )
                st.plotly_chart(fig_pit, use_container_width=True)

                # Nota metodológica
                st.caption(
                    "**MWR (Money-Weighted Return / IRR del período):** "
                    "Refleja el retorno real obtenido considerando el timing y tamaño de los flujos. "
                    "Penaliza si se invirtió más capital justo antes de un mal período. "
                    "— "
                    "**TWR Anualizado (Time-Weighted Return):** "
                    "Elimina el efecto del timing de los flujos encadenando retornos sub-trimestrales, "
                    "luego anualiza usando (1 + TWR_acum)^(365/días) − 1. "
                    "Mide la habilidad del gestor independientemente de cuándo el inversor aportó capital."
                )

    # =========================================================================
    # TAB 11 — SIMULACIÓN
    # =========================================================================
    if tab_active("🔮 Simulación"):

        # ── Cargar curvas Hamilton Lane ───────────────────────────────────────
        @st.cache_data(ttl=60)
        def load_hl_curves():
            try:
                buf = load_excel_from_drive("curvas_cobalt")
            except Exception:
                return None
            sheets = ['Buyout','Growth','Secondaries','FoF',
                      'RealEstate','CreditDistressed','CreditOther']
            curves = {}
            for s in sheets:
                try:
                    df = pd.read_excel(buf, sheet_name=s, header=None)
                    buf.seek(0)
                except Exception:
                    continue
                header_row = None
                for i, row in df.iterrows():
                    if any(str(v).strip() == 'Contributions' for v in row.values):
                        header_row = i; break
                data = df.iloc[header_row+1:].copy()
                data = data[pd.to_numeric(data.iloc[:,0], errors='coerce').notna()].copy()
                is_growth = (s == 'Growth')
                if is_growth:
                    # Growth: 0=quarter,1=contrib,2=dist,3=nav,4=unfunded
                    #         6=contrib_pct,7=rate_return,8=dist_pct,9=nav_pct
                    cdf = pd.DataFrame({
                        'quarter':       pd.to_numeric(data.iloc[:,0], errors='coerce').values,
                        'contributions': pd.to_numeric(data.iloc[:,1], errors='coerce').values,
                        'distributions': pd.to_numeric(data.iloc[:,2], errors='coerce').values,
                        'nav':           pd.to_numeric(data.iloc[:,3], errors='coerce').values,
                        'unfunded':      pd.to_numeric(data.iloc[:,4], errors='coerce').values,
                        'contrib_pct':   pd.to_numeric(data.iloc[:,6], errors='coerce').values,
                        'rate_return':   pd.to_numeric(data.iloc[:,7], errors='coerce').values,
                        'dist_pct':      pd.to_numeric(data.iloc[:,8], errors='coerce').values,
                        'nav_pct_commit':pd.to_numeric(data.iloc[:,9], errors='coerce').values,
                    })
                else:
                    # Others: 0=quarter,1=date,2=contrib,3=dist,4=nav,5=unfunded
                    #         8=contrib_pct,9=rate_return,10=dist_pct,11=nav_pct
                    cdf = pd.DataFrame({
                        'quarter':       pd.to_numeric(data.iloc[:,0], errors='coerce').values,
                        'contributions': pd.to_numeric(data.iloc[:,2], errors='coerce').values,
                        'distributions': pd.to_numeric(data.iloc[:,3], errors='coerce').values,
                        'nav':           pd.to_numeric(data.iloc[:,4], errors='coerce').values,
                        'unfunded':      pd.to_numeric(data.iloc[:,5], errors='coerce').values,
                        'contrib_pct':   pd.to_numeric(data.iloc[:,8],  errors='coerce').values,
                        'rate_return':   pd.to_numeric(data.iloc[:,9],  errors='coerce').values,
                        'dist_pct':      pd.to_numeric(data.iloc[:,10], errors='coerce').values,
                        'nav_pct_commit':pd.to_numeric(data.iloc[:,11], errors='coerce').values,
                    })
                tir_r  = df[df.iloc[:,0]=='TIR']
                tvpi_r = df[df.iloc[:,0]=='TVPI']
                curves[s] = {
                    'df':   cdf,
                    'tir':  float(tir_r.iloc[0,1])  if not tir_r.empty  else None,
                    'tvpi': float(tvpi_r.iloc[0,1]) if not tvpi_r.empty else None,
                }
            return curves

        STRAT_TO_CURVE = {
            'Buyout':         'Buyout',
            'Growth Equity':  'Growth',
            'Secondaries':    'Secondaries',
            'Venture Capital':'FoF',
            'Fund of Funds':  'FoF',
            'Single Co-Inv':  'Buyout',
            'Real Estate':    'RealEstate',
            'Credit':         'CreditOther',
        }
        CURVE_LABELS = {
            'Buyout':'Buyout','Growth':'Growth Equity','Secondaries':'Secondaries',
            'FoF':'Fund of Funds / VC','RealEstate':'Real Estate',
            'CreditDistressed':'Credit Distressed','CreditOther':'Credit Other',
        }

        def next_qend(dt):
            ts = pd.Timestamp(dt)
            mo = ts.month
            qm = ((mo-1)//3+1)*3
            return pd.Timestamp(ts.year, qm, 1) + pd.offsets.MonthEnd(0)

        def simulate_fund_hl(curve_name, commitment, current_nav,
                              current_unfunded, q_current, as_of_date, curves_dict):
            """
            Proyección usando ratios de la curva Hamilton Lane:
              Capital Call(t)  = contrib_pct(t)  × Unfunded(t-1)
              Distribuciones(t)= dist_pct(t)     × NAV(t-1)
              NAV(t)           = (NAV(t-1) - Dist(t) + Call(t)) × (1 + rate_return(t))

            Ancla: NAV y Unfunded reales del fondo en la fecha de corte.
            """
            crv    = curves_dict[curve_name]['df']
            # q_current = trimestre actual (ya transcurrido)
            # Proyectar desde q_current + 1 en adelante
            future = crv[crv['quarter'] > q_current].copy()
            if future.empty:
                return pd.DataFrame()

            # Estado inicial = valores reales del fondo hoy
            nav_t      = current_nav
            unfunded_t = current_unfunded

            # Primera fecha a proyectar = quarter end siguiente a as_of
            as_of_ts = pd.Timestamp(as_of_date)
            mo = as_of_ts.month
            qm = ((mo-1)//3+1)*3
            qdate = pd.Timestamp(as_of_ts.year, qm, 1) + pd.offsets.MonthEnd(0)
            if qdate <= as_of_ts:
                qdate += pd.offsets.QuarterEnd(1)

            rows = []
            for _, row in future.iterrows():
                contrib_pct = float(row['contrib_pct']) if pd.notna(row['contrib_pct']) else 0.0
                dist_pct    = float(row['dist_pct'])    if pd.notna(row['dist_pct'])    else 0.0
                rate_ret    = float(row['rate_return'])  if pd.notna(row['rate_return'])  else 0.0

                # Aplicar ratios al estado anterior
                call = contrib_pct * unfunded_t          # negativo (salida de caja del LP)
                dist = dist_pct    * nav_t               # positivo (entrada de caja al LP)

                # Clamp: no llamar más del unfunded disponible
                call = max(call, -unfunded_t)
                # Clamp: no distribuir más del NAV
                dist = min(dist, nav_t)

                # NAV(t) = (NAV(t-1) - Dist(t) + |Call(t)|) × (1 + RoR(t))
                # El capital llamado ENTRA al fondo → aumenta el NAV
                # Las distribuciones SALEN del fondo → reducen el NAV
                nav_new      = (nav_t - dist + abs(call)) * (1 + rate_ret)
                nav_new      = max(nav_new, 0.0)
                unfunded_new = max(unfunded_t + call, 0.0)   # call es negativo → unfunded baja

                rows.append({
                    'date':          qdate,
                    'quarter_num':   int(row['quarter']),
                    'contributions': call,
                    'distributions': dist,
                    'nav':           nav_new,
                    'unfunded':      unfunded_new,
                    'net_cf':        dist + call,
                })

                nav_t      = nav_new
                unfunded_t = unfunded_new
                qdate      = qdate + pd.offsets.QuarterEnd(1)

                # Si NAV y unfunded llegan a 0, el fondo terminó
                if nav_t <= 0 and unfunded_t <= 0:
                    break

            return pd.DataFrame(rows)

        def simulate_coinv(fund_name, commitment, current_nav,
                           exit_date, tvpi_eff, as_of_date):
            """
            Proyección lineal co-inversión:
            - NAV crece linealmente desde current_nav hasta commitment * tvpi_eff
            - En trimestre de exit: distribución total, NAV → 0
            - No hay más capital calls
            """
            as_of_ts  = pd.Timestamp(as_of_date)
            exit_ts   = pd.Timestamp(exit_date)
            exit_qend = next_qend(exit_ts)
            nav_target = commitment * tvpi_eff

            # Generar trimestres desde as_of hasta exit
            qdate = next_qend(as_of_ts)
            while qdate <= as_of_ts:
                qdate += pd.offsets.QuarterEnd(1)

            quarters = []
            q = qdate
            while q <= exit_qend + pd.offsets.QuarterEnd(1):
                quarters.append(q)
                q += pd.offsets.QuarterEnd(1)

            if not quarters:
                return pd.DataFrame()

            n = len(quarters)
            rows = []
            for i, qd in enumerate(quarters):
                is_exit = (qd >= exit_qend)
                if is_exit:
                    # Exit quarter: distribuye todo el NAV
                    nav_prev = nav_target if i > 0 else current_nav
                    rows.append({
                        'date':          qd,
                        'quarter_num':   i + 1,
                        'contributions': 0.0,
                        'distributions': nav_target,
                        'nav':           0.0,
                        'unfunded':      0.0,
                        'net_cf':        nav_target,
                    })
                    break
                else:
                    # Interpolación lineal del NAV
                    # t va de 0 (hoy) a 1 (exit)
                    t = (i + 1) / max(n - 1, 1)
                    t = min(t, 1.0)
                    nav_proj = current_nav + t * (nav_target - current_nav)
                    rows.append({
                        'date':          qd,
                        'quarter_num':   i + 1,
                        'contributions': 0.0,
                        'distributions': 0.0,
                        'nav':           nav_proj,
                        'unfunded':      0.0,
                        'net_cf':        0.0,
                    })
            return pd.DataFrame(rows)

        def agg_sim(sim_dict):
            """Agrega resultados de simulación por trimestre."""
            agg = {}
            for _, df_s in sim_dict.items():
                for _, r in df_s.iterrows():
                    d = r['date']
                    if d not in agg:
                        agg[d] = {'Capital Calls':0,'Distribuciones':0,'NAV':0,'Net Cash Flow':0}
                    agg[d]['Capital Calls']  += abs(r['contributions'])
                    agg[d]['Distribuciones'] += r['distributions']
                    agg[d]['NAV']            += r['nav']
                    agg[d]['Net Cash Flow']  += r['net_cf']
            df = pd.DataFrame.from_dict(agg, orient='index').sort_index()
            df['Capital Calls'] = -df['Capital Calls']
            df.index.name = 'Trimestre'
            return df

        def render_sim_chart_table(df_agg, title, report_curr, curr_sym, dl_key, dl_filename,
                                    nav_actual=None, unfunded_actual=None, as_of_dt=None):
            """Renderiza gráfico + tabla + botón descarga para una simulación."""
            if df_agg.empty:
                st.info("Sin datos para proyectar.")
                return

            # Métricas
            mc1,mc2,mc3,mc4 = st.columns(4)
            mc1.metric("Calls futuros", f"{curr_sym}{abs(df_agg['Capital Calls'].sum())/1e6:,.1f}M")
            mc2.metric("Distribuciones futuras", f"{curr_sym}{df_agg['Distribuciones'].sum()/1e6:,.1f}M")
            peak_nav  = df_agg['NAV'].max()
            peak_date = df_agg['NAV'].idxmax()
            mc3.metric("NAV pico", f"{curr_sym}{peak_nav/1e6:,.1f}M")
            mc4.metric("Fecha NAV pico",
                        peak_date.strftime('%d/%m/%Y') if pd.notna(peak_date) else "—")

            # Estado actual (NAV y Unfunded de partida)
            if nav_actual is not None and as_of_dt is not None:
                as_of_str = pd.Timestamp(as_of_dt).strftime('%d/%m/%Y')
                st.markdown(
                    f"<div style='background:#f0f7ff;border-left:3px solid #1a5fd4;"
                    f"padding:8px 14px;border-radius:4px;font-size:12px;margin:8px 0'>"
                    f"📌 <b>Estado en fecha de corte ({as_of_str}):</b> &nbsp;&nbsp;"
                    f"NAV actual = <b>{curr_sym}{nav_actual/1e6:,.2f}M</b>"
                    + (f" &nbsp;|&nbsp; Unfunded actual = <b>{curr_sym}{unfunded_actual/1e6:,.2f}M</b>"
                       if unfunded_actual is not None else "")
                    + "</div>",
                    unsafe_allow_html=True
                )

            st.markdown("---")

            dates_str = [d.strftime('%d-%m-%Y') for d in df_agg.index]
            fig = go.Figure()
            fig.add_trace(go.Bar(name='Capital Calls', x=dates_str,
                                  y=df_agg['Capital Calls']/1e6, marker_color='#ED7D31'))
            fig.add_trace(go.Bar(name='Distribuciones', x=dates_str,
                                  y=df_agg['Distribuciones']/1e6, marker_color='#92D050'))
            fig.add_trace(go.Scatter(name='NAV Proyectado', x=dates_str,
                                      y=df_agg['NAV']/1e6, mode='lines+markers',
                                      line=dict(color='#FFC000', width=2),
                                      marker=dict(size=4), yaxis='y2'))
            fig.add_trace(go.Scatter(name='Net Cash Flow', x=dates_str,
                                      y=df_agg['Net Cash Flow']/1e6, mode='lines',
                                      line=dict(color='#4472C4', width=1.5, dash='dot'),
                                      yaxis='y2'))
            fig.update_layout(
                barmode='relative', height=480, plot_bgcolor='white',
                title=title,
                yaxis=dict(title=f'Cash Flow ({report_curr} M)',
                           showgrid=True, gridcolor='#eef1f7', tickformat=',.0f'),
                yaxis2=dict(title=f'NAV ({report_curr} M)', overlaying='y', side='right',
                            showgrid=False, tickformat=',.0f'),
                legend=dict(orientation='h', y=1.08),
                xaxis=dict(tickangle=-45, tickfont=dict(size=9)),
                margin=dict(t=80, b=60),
            )
            st.plotly_chart(fig, use_container_width=True)

            # Tabla — con fila de estado actual al inicio
            st.markdown("#### Tabla trimestral")
            df_tbl = df_agg.copy()
            df_tbl.index = [d.strftime('%d-%m-%Y') for d in df_agg.index]
            df_tbl.index.name = 'Trimestre'

            # Insertar fila de estado actual
            if nav_actual is not None and as_of_dt is not None:
                as_of_str = pd.Timestamp(as_of_dt).strftime('%d-%m-%Y')
                estado_row = pd.DataFrame([{
                    'Capital Calls':  0.0,
                    'Distribuciones': 0.0,
                    'NAV':            nav_actual,
                    'Net Cash Flow':  0.0,
                }], index=[f'► {as_of_str} (actual)'])
                if unfunded_actual is not None:
                    estado_row['Unfunded'] = unfunded_actual
                df_tbl = pd.concat([estado_row, df_tbl])
                df_tbl.index.name = 'Trimestre'

            def color_ncf(val):
                if isinstance(val,(int,float)):
                    return f"background-color: {'#d4edda' if val>=0 else '#f8d7da'}"
                return ''

            def highlight_actual(row):
                if str(row.name).startswith('►'):
                    return ['background-color: #e8f0fe; font-weight: bold'] * len(row)
                return [''] * len(row)

            cols_fmt = {c:'{:,.0f}' for c in df_tbl.columns}
            st.dataframe(
                df_tbl.style
                      .format(cols_fmt)
                      .map(color_ncf, subset=['Net Cash Flow'])
                      .apply(highlight_actual, axis=1),
                use_container_width=True,
                height=min(420, 50+len(df_tbl)*30),
            )
            excel_download_btn(df_tbl.reset_index(), dl_filename, f"{dl_filename}.xlsx",
                               dl_filename, title, report_curr, key=f"dl_{dl_key}")

        # ── Cargar curvas ─────────────────────────────────────────────────────
        hl_curves = load_hl_curves()
        curves_ok = hl_curves is not None

        if not curves_ok:
            st.warning("⚠️ No se pudo cargar Curvas_Cobalt.xlsx desde Google Drive.")

        st.markdown("### 🔮 Simulación de Portfolio")

        # ── 3 sub-pestañas ────────────────────────────────────────────────────
        sim_tabs = st.tabs(["📈 Fondos", "🎯 Co-Inversiones", "🏛️ Portfolio Total"])

        # =====================================================================
        # SUB-TAB 1 — FONDOS (Curvas Hamilton Lane)
        # =====================================================================
        with sim_tabs[0]:
            st.markdown("#### Proyección Fondos — Curvas Hamilton Lane")
            st.caption("Cada fondo se acopla a su curva desde el trimestre actual "
                       "(basado en Fecha 1° Capital Call), usando el NAV real como ancla.")

            if not curves_ok:
                st.error("No se pudo cargar Curvas_Cobalt.xlsx desde Google Drive.")
            else:
                # Fondos hipotéticos
                st.markdown("##### ➕ Agregar fondo hipotético")
                hc1,hc2,hc3,hc4 = st.columns([2,2,1,1])
                with hc1: hypo_strat  = st.selectbox("Estrategia",
                               [s for s in STRAT_TO_CURVE if s != 'Single Co-Inv'],
                               key="hypo_strat_f")
                with hc2: hypo_comm   = st.number_input(f"Commitment ({curr_sym})",
                               min_value=100_000, value=5_000_000, step=500_000, key="hypo_comm_f")
                with hc3: hypo_yr     = st.number_input("Año inicio", 2020, 2035, 2026, key="hypo_yr_f")
                with hc4:
                    st.markdown("<br>", unsafe_allow_html=True)
                    if st.button("Agregar", key="btn_add_hypo_f"):
                        if "hypo_funds_f" not in st.session_state:
                            st.session_state["hypo_funds_f"] = []
                        st.session_state["hypo_funds_f"].append({
                            "Fund": f"Hipotético {hypo_strat[:6]} {hypo_yr}",
                            "Strategy": hypo_strat,
                            "Commitment": hypo_comm,
                        })
                        st.rerun()

                if "hypo_funds_f" not in st.session_state:
                    st.session_state["hypo_funds_f"] = []
                for i, hf in enumerate(st.session_state["hypo_funds_f"]):
                    ca,cb,cc,cd = st.columns([3,2,2,1])
                    ca.write(hf["Fund"]); cb.write(hf["Strategy"])
                    cc.write(f"{curr_sym}{hf['Commitment']/1e6:.1f}M")
                    if cd.button("❌", key=f"del_hf_{i}"):
                        st.session_state["hypo_funds_f"].pop(i); st.rerun()

                st.markdown("---")

                # Construir lista fondos (excluye co-inversiones)
                COINV_STRATS = ['Single Co-Inv']
                sim_funds_f = []

                # Fondos ya en df_final (tienen flujos)
                funds_in_final = set(df_final['Fund'].tolist())

                # Incluir también fondos de df_char_filt que aún no han iniciado
                # (no están en df_final porque no tienen flujos hasta la fecha de corte)
                df_char_sim = df_char_filt[
                    (~df_char_filt['Fund'].isin(funds_in_final)) &
                    (~df_char_filt['Strategy'].isin(COINV_STRATS))
                ].copy()

                # Combinar: filas de df_final + filas de fondos sin iniciar
                rows_to_sim = []
                for _, frow in df_final[~df_final['Strategy'].isin(COINV_STRATS)].iterrows():
                    rows_to_sim.append({
                        'Fund': frow['Fund'], 'Strategy': frow['Strategy'],
                        'Commitment': frow['Commitment'], 'NAV': frow['NAV'],
                        'Unfunded': frow['Unfunded'], 'Paid-In': frow['Paid-In'],
                        'from_final': True,
                    })
                for _, crow in df_char_sim.iterrows():
                    strat = crow.get('Strategy', '')
                    if not STRAT_TO_CURVE.get(strat): continue
                    fx_c = fx_map.get(crow['Fecha Commitment'].date(), fx_today)
                    comm_rep = convert_amount(crow['Commitment'], crow['Currency'], report_curr, fx_c)
                    rows_to_sim.append({
                        'Fund': crow['Fund'], 'Strategy': strat,
                        'Commitment': comm_rep, 'NAV': 0.0,
                        'Unfunded': comm_rep, 'Paid-In': 0.0,
                        'from_final': False,
                    })

                for frow in rows_to_sim:
                    strat = frow['Strategy']
                    curve_name = STRAT_TO_CURVE.get(strat)
                    if not curve_name: continue
                    fund_name        = frow['Fund']
                    commitment       = frow['Commitment']
                    current_nav      = frow['NAV']
                    current_unfunded = frow['Unfunded']

                    # Saltar fondos terminados: NAV=0 pero tienen Paid-In
                    if current_nav <= 0 and frow['Paid-In'] > 0:
                        continue

                    char_row = df_char[df_char['Fund'] == fund_name]
                    if char_row.empty: continue
                    first_call_date = char_row['Fecha 1er Call'].iloc[0]
                    if pd.isna(first_call_date): continue
                    first_call_ts   = pd.Timestamp(first_call_date)
                    mo = first_call_ts.month
                    qm = ((mo-1)//3+1)*3
                    first_call_qend = pd.Timestamp(first_call_ts.year, qm, 1) + pd.offsets.MonthEnd(0)
                    as_of_ts    = pd.Timestamp(as_of_date_dt)
                    months_diff = (as_of_ts.year - first_call_qend.year)*12 \
                                  + (as_of_ts.month - first_call_qend.month)

                    if months_diff < 0:
                        # Fondo aún no ha hecho primer call — proyectar desde Q1
                        q_current = 0
                        current_nav      = 0.0
                        current_unfunded = commitment
                    else:
                        q_current = len(pd.date_range(first_call_qend, as_of_ts, freq='QE'))
                        crv_len   = len(hl_curves[curve_name]['df'])
                        q_current = min(q_current, crv_len - 1)

                    sim_funds_f.append({
                        'name': fund_name, 'curve': curve_name,
                        'commitment': commitment, 'nav': current_nav,
                        'unfunded': current_unfunded,
                        'q_current': q_current, 'strategy': strat,
                        'first_call': first_call_qend,
                    })

                # Hipotéticos
                for hf in st.session_state["hypo_funds_f"]:
                    curve_name = STRAT_TO_CURVE.get(hf['Strategy'])
                    if not curve_name: continue
                    comm_rep = convert_amount(hf['Commitment'], 'USD', report_curr, fx_today)
                    sim_funds_f.append({
                        'name': hf['Fund'], 'curve': curve_name,
                        'commitment': comm_rep, 'nav': 0.0,
                        'unfunded': comm_rep,   # hipotético: todo unfunded
                        'q_current': 1, 'strategy': hf['Strategy'],
                        'first_call': pd.Timestamp(as_of_date_dt),
                    })

                # Simular
                all_sim_f = {}
                for sf in sim_funds_f:
                    df_s = simulate_fund_hl(
                        sf['curve'], sf['commitment'], sf['nav'],
                        sf['unfunded'], sf['q_current'], as_of_date_dt, hl_curves
                    )
                    if not df_s.empty:
                        all_sim_f[sf['name']] = df_s

                df_agg_f = agg_sim(all_sim_f)
                # Totales actuales para mostrar en tabla
                total_nav_f      = sum(sf['nav']      for sf in sim_funds_f)
                total_unfunded_f = sum(sf['unfunded'] for sf in sim_funds_f)
                render_sim_chart_table(df_agg_f,
                    f"Proyección Fondos — Curvas Hamilton Lane ({report_curr} M)",
                    report_curr, curr_sym, "sim_fondos", "simulacion_fondos",
                    nav_actual=total_nav_f,
                    unfunded_actual=total_unfunded_f,
                    as_of_dt=as_of_date_dt)

                # Detalle por fondo
                with st.expander("📋 Ver detalle por fondo", expanded=False):
                    for sf in sim_funds_f:
                        if sf['name'] not in all_sim_f: continue
                        df_fd = all_sim_f[sf['name']]
                        fc_str = sf['first_call'].strftime('%d/%m/%Y') \
                                 if hasattr(sf['first_call'],'strftime') else '—'
                        as_of_str = pd.Timestamp(as_of_date_dt).strftime('%d/%m/%Y')
                        st.markdown(f"**{sf['name']}** — {sf['strategy']} "
                                    f"({CURVE_LABELS.get(sf['curve'],sf['curve'])}) "
                                    f"| Commitment: {curr_sym}{sf['commitment']/1e6:.2f}M "
                                    f"| 1er Call: {fc_str} "
                                    f"| Trim. curva actual: **{'Aún no iniciado' if sf['q_current'] == 0 else sf['q_current']}**")
                        # Fila estado actual
                        as_of_label = f"► {pd.Timestamp(as_of_date_dt).strftime('%d-%m-%Y')} (actual)"
                        estado_fondo = pd.DataFrame([{
                            'Fecha':         as_of_str,
                            'Contributions': 0.0,
                            'Distributions': 0.0,
                            'NAV':           sf['nav'],
                            'Unfunded':      sf['unfunded'],
                            'Net CF':        0.0,
                        }], index=[as_of_label])
                        estado_fondo.index.name = 'Trim. Curva'

                        df_show = df_fd[['quarter_num','date','contributions',
                                         'distributions','nav','unfunded','net_cf']].copy()
                        df_show['date'] = df_show['date'].dt.strftime('%d-%m-%Y')
                        df_show.columns = ['Trim. Curva','Fecha','Contributions',
                                            'Distributions','NAV','Unfunded','Net CF']
                        df_show = df_show.set_index('Trim. Curva')

                        # Combinar fila actual + proyección
                        df_show_full = pd.concat([estado_fondo, df_show])

                        def hi_actual_row(row):
                            if str(row.name).startswith('►'):
                                return ['background-color:#e8f0fe;font-weight:bold'] * len(row)
                            return [''] * len(row)

                        st.dataframe(
                            df_show_full.style
                                .format({c:'{:,.0f}' for c in df_show_full.columns if c!='Fecha'})
                                .apply(hi_actual_row, axis=1),
                            use_container_width=True,
                            height=min(320, 50+len(df_show_full)*28))

        # =====================================================================
        # SUB-TAB 2 — CO-INVERSIONES
        # =====================================================================
        with sim_tabs[1]:
            st.markdown("#### Proyección Co-Inversiones — Modelo Lineal")

            coinv_ok = not df_coinv.empty
            if not coinv_ok:
                st.warning("No se encontró la pestaña `Characteristics_CoInv` en `datos.xlsx`. "
                           "Crea esa pestaña con las columnas: Fund, GP, Strategy, Currency, "
                           "Vintage, Commitment, Fecha Commitment, Geography, "
                           "Fecha 1° Capital Call, UW Exit Date, UW TVPI, "
                           "New Exit Date *(opcional)*, New TVPI *(opcional)*.")
            else:
                # Mostrar tabla de co-inversiones cargadas
                cols_show = ['Fund','GP','Commitment','Fecha 1° Capital Call',
                             'UW Exit Date','UW TVPI','New Exit Date','New TVPI']
                cols_show = [c for c in cols_show if c in df_coinv.columns]
                st.markdown("##### Co-inversiones cargadas")
                st.dataframe(df_coinv[cols_show].style.format(
                    {c:'{:,.0f}' for c in ['Commitment'] if c in cols_show}),
                    use_container_width=True, height=min(300, 50+len(df_coinv)*32))

                st.markdown("---")

                # Simular cada co-inversión
                all_sim_ci = {}
                skipped    = []

                for _, ci in df_coinv.iterrows():
                    fund_name  = str(ci.get('Fund','')).strip()
                    currency   = str(ci.get('Currency','USD')).strip().upper()

                    # Commitment en moneda de reporte
                    comm_raw = float(ci.get('Commitment', 0) or 0)
                    commitment = convert_amount(comm_raw, currency, report_curr, fx_today)

                    # NAV actual desde df_final (cashflows reales)
                    nav_row = df_final[df_final['Fund'] == fund_name]
                    current_nav = float(nav_row['NAV'].iloc[0]) if not nav_row.empty else 0.0

                    # Exit date: New ?? UW
                    exit_date = ci.get('New Exit Date') \
                                if pd.notna(ci.get('New Exit Date')) \
                                else ci.get('UW Exit Date')
                    # TVPI: New ?? UW
                    tvpi_eff  = float(ci.get('New TVPI'))  \
                                if pd.notna(ci.get('New TVPI')) \
                                else float(ci.get('UW TVPI', 1.5) or 1.5)

                    if pd.isna(exit_date) or commitment <= 0:
                        skipped.append(fund_name)
                        continue

                    # Solo proyectar si el exit es futuro
                    if pd.Timestamp(exit_date) <= pd.Timestamp(as_of_date_dt):
                        skipped.append(f"{fund_name} (ya realizó exit)")
                        continue

                    df_ci = simulate_coinv(fund_name, commitment, current_nav,
                                           exit_date, tvpi_eff, as_of_date_dt)
                    if not df_ci.empty:
                        all_sim_ci[fund_name] = df_ci

                if skipped:
                    st.caption(f"⚠️ Omitidos (sin datos o exit pasado): {', '.join(skipped)}")

                df_agg_ci = agg_sim(all_sim_ci)
                # Totales NAV actuales de co-inversiones
                total_nav_ci = sum(
                    float(df_final[df_final['Fund']==str(ci.get('Fund','')).strip()]['NAV'].iloc[0])
                    if not df_final[df_final['Fund']==str(ci.get('Fund','')).strip()].empty else 0.0
                    for _, ci in df_coinv.iterrows()
                    if str(ci.get('Fund','')).strip() in all_sim_ci
                )
                render_sim_chart_table(df_agg_ci,
                    f"Proyección Co-Inversiones — Modelo Lineal ({report_curr} M)",
                    report_curr, curr_sym, "sim_coinv", "simulacion_coinversiones",
                    nav_actual=total_nav_ci,
                    as_of_dt=as_of_date_dt)

                # Detalle por co-inversión
                with st.expander("📋 Ver detalle por co-inversión", expanded=False):
                    for _, ci in df_coinv.iterrows():
                        fname = str(ci.get('Fund','')).strip()
                        if fname not in all_sim_ci: continue
                        currency  = str(ci.get('Currency','USD')).strip().upper()
                        comm_raw  = float(ci.get('Commitment',0) or 0)
                        commitment= convert_amount(comm_raw, currency, report_curr, fx_today)
                        nav_row   = df_final[df_final['Fund']==fname]
                        cur_nav   = float(nav_row['NAV'].iloc[0]) if not nav_row.empty else 0.0
                        exit_date = ci.get('New Exit Date') \
                                    if pd.notna(ci.get('New Exit Date')) \
                                    else ci.get('UW Exit Date')
                        tvpi_eff  = float(ci.get('New TVPI')) \
                                    if pd.notna(ci.get('New TVPI')) \
                                    else float(ci.get('UW TVPI',1.5) or 1.5)
                        st.markdown(
                            f"**{fname}** "
                            f"| Commitment: {curr_sym}{commitment/1e6:.2f}M "
                            f"| NAV actual: {curr_sym}{cur_nav/1e6:.2f}M "
                            f"| Exit: {pd.Timestamp(exit_date).strftime('%d/%m/%Y')} "
                            f"| TVPI efectivo: **{tvpi_eff:.2f}x** "
                            f"| NAV objetivo: {curr_sym}{commitment*tvpi_eff/1e6:.2f}M"
                        )
                        df_fd = all_sim_ci[fname].copy()
                        df_fd['date'] = df_fd['date'].dt.strftime('%d-%m-%Y')
                        df_fd = df_fd.rename(columns={
                            'quarter_num':'Trimestre','date':'Fecha',
                            'contributions':'Contributions','distributions':'Distributions',
                            'nav':'NAV','unfunded':'Unfunded','net_cf':'Net CF'
                        }).set_index('Trimestre')
                        st.dataframe(df_fd.style.format(
                            {c:'{:,.0f}' for c in df_fd.columns if c!='Fecha'}),
                            use_container_width=True,
                            height=min(300, 50+len(df_fd)*28))

        # =====================================================================
        # SUB-TAB 3 — PORTFOLIO TOTAL
        # =====================================================================
        with sim_tabs[2]:
            st.markdown("#### Proyección Portfolio Total")
            st.caption("Suma de fondos (curvas Hamilton Lane) + co-inversiones (modelo lineal).")

            # Combinar ambos diccionarios de simulaciones
            all_sim_total = {}
            if curves_ok and 'all_sim_f' in dir():
                all_sim_total.update(all_sim_f)
            if coinv_ok and 'all_sim_ci' in dir():
                all_sim_total.update(all_sim_ci)

            df_agg_total = agg_sim(all_sim_total)

            if df_agg_total.empty:
                st.info("Ejecuta primero las pestañas Fondos y Co-Inversiones para ver el total.")
            else:
                # NAV y Unfunded totales del portfolio
                nav_total_actual      = t_nav   # ya calculado en el scope principal
                unfunded_total_actual = t_comm - t_paid
                render_sim_chart_table(df_agg_total,
                    f"Proyección Portfolio Total ({report_curr} M)",
                    report_curr, curr_sym, "sim_total", "simulacion_portfolio_total",
                    nav_actual=nav_total_actual,
                    unfunded_actual=unfunded_total_actual,
                    as_of_dt=as_of_date_dt)

                # Desglose fondos vs co-inversiones en el mismo gráfico
                st.markdown("#### Desglose NAV: Fondos vs Co-Inversiones")
                df_f_agg  = agg_sim(all_sim_f)  if 'all_sim_f' in dir()  and all_sim_f  else pd.DataFrame()
                df_ci_agg = agg_sim(all_sim_ci) if 'all_sim_ci' in dir() and all_sim_ci else pd.DataFrame()

                all_dates = sorted(set(
                    (list(df_f_agg.index)  if not df_f_agg.empty  else []) +
                    (list(df_ci_agg.index) if not df_ci_agg.empty else [])
                ))
                if all_dates:
                    nav_f  = [df_f_agg.loc[d,'NAV']  if d in df_f_agg.index  else 0 for d in all_dates]
                    nav_ci = [df_ci_agg.loc[d,'NAV'] if d in df_ci_agg.index else 0 for d in all_dates]
                    dates_str = [d.strftime('%d-%m-%Y') for d in all_dates]

                    fig_split = go.Figure()
                    fig_split.add_trace(go.Bar(name='NAV Fondos', x=dates_str,
                                                y=[v/1e6 for v in nav_f],
                                                marker_color='#002060'))
                    fig_split.add_trace(go.Bar(name='NAV Co-Inversiones', x=dates_str,
                                                y=[v/1e6 for v in nav_ci],
                                                marker_color='#4472C4'))
                    fig_split.update_layout(
                        barmode='stack', height=400, plot_bgcolor='white',
                        title=f'NAV Proyectado: Fondos vs Co-Inversiones ({report_curr} M)',
                        yaxis=dict(title=f'{report_curr} M', showgrid=True,
                                   gridcolor='#eef1f7', tickformat=',.0f'),
                        legend=dict(orientation='h', y=1.08),
                        xaxis=dict(tickangle=-45, tickfont=dict(size=9)),
                        margin=dict(t=80, b=60),
                    )
                    st.plotly_chart(fig_split, use_container_width=True)

except Exception as e:
    st.error(f"Error detectado: {e}")
